"""
Bulk Decision Maker Lookup Script
Find decision makers at companies using BlitzAPI.

Input: Excel/CSV with Company Name and/or Website columns
Output: Excel with found decision makers (name, title, email, LinkedIn)

Cost: ~3-5 credits per company (domain lookup + people + emails)
"""

import os
import sys
import csv
import time
import argparse
from pathlib import Path
from typing import List, Dict, Optional
from dotenv import load_dotenv

load_dotenv()

try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import BlitzAPI
try:
    from blitz_api import BlitzAPI, BlitzAPIError
    BLITZ_AVAILABLE = True
except ImportError:
    BLITZ_AVAILABLE = False


# Configuration
DEFAULT_MAX_RESULTS = 2
API_DELAY = 0.5  # Seconds between API calls to avoid rate limiting


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns (case-insensitive)."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx
    return None


def clean_domain(url: str) -> str:
    """Extract clean domain from URL or domain string."""
    if not url:
        return ""

    url = str(url).strip().lower()

    # Remove protocol
    for prefix in ['https://', 'http://', 'www.']:
        if url.startswith(prefix):
            url = url[len(prefix):]

    # Remove path
    url = url.split('/')[0]

    # Remove port
    url = url.split(':')[0]

    return url


def process_companies(
    file_path: str,
    max_results: int = DEFAULT_MAX_RESULTS,
    with_email: bool = True,
    verbose: bool = True
) -> Dict:
    """
    Process Excel/CSV file and lookup decision makers for each company.

    Args:
        file_path: Path to input file
        max_results: Max decision makers per company
        with_email: Whether to enrich with emails
        verbose: Print progress

    Returns:
        Dict with processing statistics
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    if not BLITZ_AVAILABLE:
        return {"error": "BlitzAPI not available. Check blitz_api.py exists."}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    # Initialize BlitzAPI
    try:
        api = BlitzAPI()
        credits_before = api.get_key_info().remaining_credits
        if verbose:
            print(f"BlitzAPI connected. Credits available: {credits_before}")
    except BlitzAPIError as e:
        return {"error": f"BlitzAPI error: {e.message}"}

    # Load input file
    if file_path.suffix.lower() == '.csv':
        # Read CSV
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
    else:
        # Read Excel
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            data_rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
            wb.close()
        except Exception as e:
            return {"error": f"Failed to load file: {e}"}

    # Find required columns
    company_col = find_column_index(headers, ['company', 'company_name', 'company name', 'account', 'organization'])
    website_col = find_column_index(headers, ['website', 'web', 'domain', 'url', 'web address', 'company_website'])

    if company_col is None and website_col is None:
        return {"error": "No Company or Website column found. Need at least one."}

    if verbose:
        print(f"\nDetected columns:")
        if company_col is not None:
            print(f"  Company: column {company_col + 1}")
        if website_col is not None:
            print(f"  Website: column {website_col + 1}")
        print(f"\nTotal companies: {len(data_rows)}")
        print(f"Max results per company: {max_results}")
        print(f"Email enrichment: {'Yes' if with_email else 'No'}")

    # Collect unique companies with websites
    companies_to_process = []
    seen_domains = set()

    for row in data_rows:
        company_name = row[company_col] if company_col is not None and company_col < len(row) else None
        website = row[website_col] if website_col is not None and website_col < len(row) else None

        domain = clean_domain(website) if website else None

        # Skip if no domain (BlitzAPI requires domain or LinkedIn URL)
        if not domain:
            continue

        # Skip duplicates
        if domain in seen_domains:
            continue
        seen_domains.add(domain)

        companies_to_process.append({
            'company_name': str(company_name).strip() if company_name else None,
            'domain': domain
        })

    if verbose:
        print(f"Companies with valid domains: {len(companies_to_process)}")
        skipped = len(data_rows) - len(companies_to_process)
        if skipped > 0:
            print(f"Skipped (no domain or duplicate): {skipped}")

    if not companies_to_process:
        return {"error": "No companies with valid domains found. BlitzAPI requires a website/domain."}

    # Process each company
    all_results = []
    stats = {
        "total_companies": len(companies_to_process),
        "processed": 0,
        "found": 0,
        "not_found": 0,
        "errors": 0,
        "decision_makers_found": 0
    }

    if verbose:
        print(f"\nProcessing {len(companies_to_process)} companies...\n")

    for idx, company in enumerate(companies_to_process, 1):
        if verbose:
            print(f"  [{idx}/{len(companies_to_process)}] {company['company_name'] or company['domain']}...", end=" ")

        try:
            # Search for decision makers
            results = api.search_decision_makers(
                company_domain=company['domain'],
                with_email=with_email
            )

            if results:
                stats['found'] += 1
                stats['decision_makers_found'] += len(results)

                for person in results[:max_results]:
                    all_results.append({
                        'Company Name': company['company_name'] or '',
                        'Company Domain': company['domain'],
                        'First Name': person.get('first_name', ''),
                        'Last Name': person.get('last_name', ''),
                        'Full Name': person.get('full_name', ''),
                        'Title': person.get('title', ''),
                        'Email': person.get('email', ''),
                        'Email Found': 'Yes' if person.get('email_found') else 'No',
                        'LinkedIn URL': person.get('linkedin_url', ''),
                        'Location': person.get('location', ''),
                        'ICP Rank': person.get('icp_rank', ''),
                        'What Matched': ', '.join(person.get('what_matched', []))
                    })

                if verbose:
                    print(f"Found {len(results)} decision maker(s)")
            else:
                stats['not_found'] += 1
                if verbose:
                    print("No decision makers found")

        except BlitzAPIError as e:
            stats['errors'] += 1
            if verbose:
                print(f"Error: {e.message}")

            # Check if out of credits
            if e.status_code == 402:
                print("\n⚠️  Out of credits! Stopping.")
                break

        except Exception as e:
            stats['errors'] += 1
            if verbose:
                print(f"Error: {str(e)}")

        stats['processed'] += 1

        # Rate limiting
        time.sleep(API_DELAY)

    # Get credits used
    try:
        credits_after = api.get_key_info().remaining_credits
        stats['credits_used'] = credits_before - credits_after
    except:
        stats['credits_used'] = 'Unknown'

    # Save results
    if all_results:
        output_path = file_path.parent / f"{file_path.stem}_DECISION_MAKERS_FOUND.xlsx"

        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Decision Makers"

        # Write headers
        result_headers = list(all_results[0].keys())
        for col, header in enumerate(result_headers, 1):
            ws_out.cell(row=1, column=col, value=header)

        # Write data
        for row_idx, result in enumerate(all_results, 2):
            for col, header in enumerate(result_headers, 1):
                ws_out.cell(row=row_idx, column=col, value=result[header])

        wb_out.save(output_path)
        stats['output_file'] = str(output_path)

        if verbose:
            print(f"\nResults saved to: {output_path}")
    else:
        stats['output_file'] = None
        if verbose:
            print("\nNo results to save.")

    return stats


def main():
    parser = argparse.ArgumentParser(
        description='Bulk lookup decision makers at companies using BlitzAPI'
    )
    parser.add_argument('file', help='Path to Excel/CSV file with companies')
    parser.add_argument(
        '--max-results',
        type=int,
        default=DEFAULT_MAX_RESULTS,
        help=f'Max decision makers per company (default: {DEFAULT_MAX_RESULTS})'
    )
    parser.add_argument(
        '--no-email',
        action='store_true',
        help='Skip email enrichment (saves credits)'
    )

    args = parser.parse_args()

    print(f"\n{'='*60}")
    print("BULK DECISION MAKER LOOKUP")
    print(f"{'='*60}")
    print(f"File: {args.file}")
    print(f"Max results per company: {args.max_results}")
    print(f"Email enrichment: {'No' if args.no_email else 'Yes'}")

    start_time = time.time()

    result = process_companies(
        args.file,
        max_results=args.max_results,
        with_email=not args.no_email,
        verbose=True
    )

    elapsed = time.time() - start_time

    print(f"\n{'='*60}")
    print("RESULTS")
    print(f"{'='*60}")

    if "error" in result:
        print(f"Error: {result['error']}")
        sys.exit(1)

    print(f"Companies processed: {result['processed']}/{result['total_companies']}")
    print(f"Companies with DMs found: {result['found']}")
    print(f"Companies with no DMs: {result['not_found']}")
    print(f"Errors: {result['errors']}")
    print(f"Total decision makers found: {result['decision_makers_found']}")
    print(f"Credits used: {result['credits_used']}")
    print(f"Time elapsed: {elapsed:.1f} seconds")

    if result.get('output_file'):
        print(f"\nOutput saved to: {result['output_file']}")


if __name__ == '__main__':
    main()
