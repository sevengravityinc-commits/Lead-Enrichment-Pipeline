"""
Batch Categorize Niche Script
Fast batch processing to categorize companies into Advertising, Marketing, or PR.
Uses OpenRouter API with GPT-4o-mini for cost-effective classification.
"""

import os
import sys
import json
import time
import requests
from pathlib import Path
from typing import List, Dict, Optional
from collections import Counter
from dotenv import load_dotenv

load_dotenv()

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Configuration
DEFAULT_BATCH_SIZE = 20  # Companies per API call (≤1000 records)
LARGE_FILE_BATCH_SIZE = 50  # Companies per API call (>1000 records)
LARGE_FILE_THRESHOLD = 1000  # Switch to larger batch size above this row count
API_DELAY = 0.5  # Seconds between API calls (rate limiting)
OPENROUTER_API_KEY = os.getenv('OPENROUTER_API_KEY')
MODEL = "openai/gpt-4o-mini"  # Fast, cheap, accurate for classification


CLASSIFICATION_PROMPT = """You are classifying companies for targeted email campaigns.

Classify each company into exactly ONE category based on their primary business focus:
- **Advertising** - ad agencies, media buying, ad tech, creative agencies, billboard/outdoor advertising
- **Marketing** - digital marketing, content marketing, brand strategy, SEO, social media marketing, market research
- **PR** - public relations, communications, crisis management, media relations, corporate communications

Rules:
1. Every company gets exactly ONE category
2. Pick the PRIMARY focus if company spans multiple areas
3. Use the company name and job title as clues
4. When ambiguous, use these guidelines:
   - "Agency" without specifics → Marketing
   - "Communications" → PR
   - "Media" alone → Advertising
   - "Creative" → Advertising
   - "Digital" → Marketing
   - "Branding" → Marketing

OUTPUT FORMAT: Return ONLY a JSON array of categories in the same order as input. No explanations.
Example: ["Marketing", "PR", "Advertising", "Marketing"]"""


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx
    return None


def batch_categorize(companies: List[Dict]) -> List[str]:
    """
    Send a batch of companies to OpenRouter and get categories back.

    Args:
        companies: List of dicts with 'name' and 'title' keys

    Returns:
        List of categories (Advertising, Marketing, or PR)
    """
    # Format companies for prompt
    company_lines = []
    for c in companies:
        name = c.get('name', 'Unknown')
        title = c.get('title', '')
        if title:
            company_lines.append(f"{name} | {title}")
        else:
            company_lines.append(f"{name}")

    companies_json = json.dumps(company_lines)

    try:
        response = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": MODEL,
                "max_tokens": len(companies) * 20,  # ~20 tokens per category
                "messages": [
                    {"role": "system", "content": CLASSIFICATION_PROMPT},
                    {"role": "user", "content": f"Classify these companies:\n{companies_json}"}
                ]
            },
            timeout=60
        )

        response.raise_for_status()
        data = response.json()
        response_text = data['choices'][0]['message']['content'].strip()

        # Parse JSON response
        # Handle potential markdown code blocks
        if response_text.startswith("```"):
            response_text = response_text.split("```")[1]
            if response_text.startswith("json"):
                response_text = response_text[4:]
            response_text = response_text.strip()

        categories = json.loads(response_text)

        # Validate categories
        valid_categories = {"Advertising", "Marketing", "PR"}
        validated = []
        for cat in categories:
            if cat in valid_categories:
                validated.append(cat)
            else:
                # Default to Marketing for invalid responses
                validated.append("Marketing")

        # Ensure we have the right number of results
        while len(validated) < len(companies):
            validated.append("Marketing")

        return validated[:len(companies)]

    except json.JSONDecodeError as e:
        print(f"  Warning: JSON parse error, using default. Error: {e}")
        return ["Marketing"] * len(companies)
    except Exception as e:
        print(f"  Warning: API error, using default. Error: {e}")
        return ["Marketing"] * len(companies)


def process_excel_file(file_path: str, verbose: bool = True) -> Dict:
    """
    Process Excel file and add Verified_Niche column using batch categorization.

    Args:
        file_path: Path to Excel file
        verbose: Whether to print progress

    Returns:
        Dict with processing statistics
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    if not OPENROUTER_API_KEY:
        return {"error": "OPENROUTER_API_KEY not set in environment"}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    # Load workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {"error": f"Failed to load Excel file: {e}"}

    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Find required columns
    company_col = find_column_index(headers, ['company', 'company_name', 'company name', 'name'])
    title_col = find_column_index(headers, ['title', 'job_title', 'job title', 'position'])

    if company_col is None:
        return {"error": "Could not find Company Name column"}

    # Find or create Verified_Niche column
    niche_col = find_column_index(headers, ['verified_niche', 'verified niche'])
    if niche_col is None:
        niche_col = len(headers)
        ws.cell(row=1, column=niche_col + 1, value='Verified_Niche')

    # Get total rows
    total_rows = ws.max_row - 1  # Exclude header

    # Dynamic batch sizing: use larger batches for files with >1000 records
    if total_rows > LARGE_FILE_THRESHOLD:
        batch_size = LARGE_FILE_BATCH_SIZE
    else:
        batch_size = DEFAULT_BATCH_SIZE

    if verbose:
        print(f"Processing {total_rows} rows in batches of {batch_size}...")
        if total_rows > LARGE_FILE_THRESHOLD:
            print(f"  (Using larger batch size for {total_rows} records)")
        print(f"Estimated API calls: {(total_rows + batch_size - 1) // batch_size}")
        print(f"Model: {MODEL}")
        print()

    # Statistics
    stats = {
        "total_rows": total_rows,
        "processed": 0,
        "skipped": 0,
        "batches": 0,
        "distribution": Counter()
    }

    # Process in batches
    batch = []
    batch_rows = []  # Track which Excel rows are in current batch

    for row_idx in range(2, ws.max_row + 1):
        # Check if already categorized
        existing = ws.cell(row=row_idx, column=niche_col + 1).value
        if existing and existing.strip() in ["Advertising", "Marketing", "PR"]:
            stats["skipped"] += 1
            stats["distribution"][existing.strip()] += 1
            continue

        # Get company data
        company_name = ws.cell(row=row_idx, column=company_col + 1).value
        title = ws.cell(row=row_idx, column=title_col + 1).value if title_col is not None else None

        if not company_name:
            ws.cell(row=row_idx, column=niche_col + 1, value='Marketing')  # Default
            stats["processed"] += 1
            stats["distribution"]["Marketing"] += 1
            continue

        batch.append({
            "name": str(company_name).strip(),
            "title": str(title).strip() if title else ""
        })
        batch_rows.append(row_idx)

        # Process batch when full
        if len(batch) >= batch_size:
            stats["batches"] += 1

            if verbose:
                progress = (stats["processed"] + stats["skipped"]) / total_rows * 100
                print(f"  Batch {stats['batches']}: rows {batch_rows[0]}-{batch_rows[-1]} ({progress:.1f}% complete)")

            # Get categories from AI
            categories = batch_categorize(batch)

            # Write results
            for row, category in zip(batch_rows, categories):
                ws.cell(row=row, column=niche_col + 1, value=category)
                stats["distribution"][category] += 1
                stats["processed"] += 1

            # Save progress
            try:
                wb.save(file_path)
            except Exception as e:
                print(f"  Warning: Could not save progress: {e}")

            # Clear batch
            batch = []
            batch_rows = []

            # Rate limiting
            time.sleep(API_DELAY)

    # Process remaining batch
    if batch:
        stats["batches"] += 1

        if verbose:
            print(f"  Batch {stats['batches']} (final): rows {batch_rows[0]}-{batch_rows[-1]}")

        categories = batch_categorize(batch)

        for row, category in zip(batch_rows, categories):
            ws.cell(row=row, column=niche_col + 1, value=category)
            stats["distribution"][category] += 1
            stats["processed"] += 1

    # Final save
    try:
        wb.save(file_path)
        if verbose:
            print(f"\nSaved: {file_path}")
    except Exception as e:
        return {"error": f"Failed to save: {e}", **stats}

    wb.close()

    return stats


def main():
    if len(sys.argv) < 2:
        print("Usage: python batch_categorize_niche.py <excel_file>")
        print("       Adds Verified_Niche column with: Advertising, Marketing, or PR")
        sys.exit(1)

    file_path = sys.argv[1]

    print(f"\n{'='*60}")
    print("BATCH NICHE CATEGORIZATION")
    print(f"{'='*60}")
    print(f"File: {file_path}")
    print(f"Batch size: dynamic (20 for ≤1000 rows, 50 for >1000 rows)")
    print(f"API: OpenRouter ({MODEL})")
    print()

    start_time = time.time()

    result = process_excel_file(file_path, verbose=True)

    elapsed = time.time() - start_time

    print(f"\n{'='*60}")
    print("RESULTS")
    print(f"{'='*60}")

    if "error" in result:
        print(f"Error: {result['error']}")
        sys.exit(1)

    print(f"Total rows: {result['total_rows']}")
    print(f"Processed: {result['processed']}")
    print(f"Skipped (already done): {result['skipped']}")
    print(f"API batches: {result['batches']}")
    print(f"Time elapsed: {elapsed:.1f} seconds")
    print()
    print("Category Distribution:")
    for category, count in sorted(result['distribution'].items()):
        pct = count / result['total_rows'] * 100 if result['total_rows'] > 0 else 0
        print(f"  {category}: {count} ({pct:.1f}%)")


if __name__ == '__main__':
    main()
