"""
Add Bounce Ban Verification Columns Script
Processes Excel files and adds email verification status columns using Bounce Ban API.
Specializes in catch-all email verification (second pass after Million Verifier).
"""

import os
import sys
import shutil
import time
from pathlib import Path
from typing import Optional, List, Dict
from collections import Counter

try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from bounceban_api import verify_emails, get_task_status, download_results


# Column name patterns for finding email
EMAIL_COLUMN_PATTERNS = [
    'email', 'e-mail', 'email address', 'work email', 'business email',
    'contact email', 'primary email', 'mail'
]

# Campaign-relevant columns for sendable export (exclude verification columns)
CAMPAIGN_COLUMNS = [
    'email', 'e-mail', 'email address',
    'first name', 'firstname', 'first_name',
    'last name', 'lastname', 'last_name',
    'company', 'company_name', 'organization',
    'clean_company_name', 'clean company name',
    'niche'
]


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx

    return None


def get_headers(ws) -> List[str]:
    """Get header row from worksheet."""
    return [cell.value for cell in ws[1]]


def check_duplicate_verification(ws, headers: List[str], force: bool = False) -> bool:
    """
    Check if Bounce Ban verification has already been run on this file.

    Args:
        ws: Worksheet to check
        headers: Header row
        force: If True, allow overwrite

    Returns:
        True if safe to proceed, False if duplicate detected and not forced
    """
    bb_result_col_idx = find_column_index(headers, ['bb_result', 'bounceban_result'])

    if bb_result_col_idx is None:
        return True  # No BB columns yet, safe to proceed

    # Check if column has data (more than just header)
    has_data = False
    for row_idx in range(2, min(ws.max_row + 1, 12)):  # Check first 10 rows
        cell_value = ws.cell(row=row_idx, column=bb_result_col_idx + 1).value
        if cell_value and str(cell_value).strip():
            has_data = True
            break

    if not has_data:
        return True  # Column exists but empty, safe to proceed

    if force:
        print("WARNING: Overwriting existing Bounce Ban verification (--force enabled)")
        time.sleep(3)  # 3-second warning
        return True

    return False  # Duplicate detected, not forced


def export_sendable_file(wb, file_path: Path, bb_sendable_col_idx: int, headers: List[str], verbose: bool = True) -> Dict:
    """
    Export rows where BB_Sendable = TRUE to a separate file for SmartLead.
    Excludes all verification columns (MV_*, BB_*).

    Args:
        wb: Loaded workbook
        file_path: Original file path
        bb_sendable_col_idx: Index of BB_Sendable column
        headers: Original headers
        verbose: Print progress

    Returns:
        Dict with export statistics
    """
    ws = wb.active

    # Identify campaign-relevant columns (exclude verification columns)
    campaign_col_indices = []
    campaign_headers = []

    for idx, header in enumerate(headers):
        if not header:
            continue
        header_lower = header.lower().strip()

        # Skip verification columns
        if header_lower.startswith('mv_') or header_lower.startswith('bb_'):
            continue

        # Include if matches campaign patterns
        if any(pattern in header_lower for pattern in CAMPAIGN_COLUMNS):
            campaign_col_indices.append(idx)
            campaign_headers.append(header)

    if not campaign_col_indices:
        return {"error": "No campaign columns found to export"}

    # Create new workbook for sendable emails
    sendable_wb = Workbook()
    sendable_ws = sendable_wb.active
    sendable_ws.title = "Sendable Leads"

    # Write headers
    for col_idx, header in enumerate(campaign_headers):
        sendable_ws.cell(row=1, column=col_idx + 1, value=header)

    # Copy rows where BB_Sendable = TRUE
    sendable_count = 0
    for row_idx in range(2, ws.max_row + 1):
        sendable_value = ws.cell(row=row_idx, column=bb_sendable_col_idx + 1).value

        if sendable_value is True or str(sendable_value).upper() == 'TRUE':
            sendable_count += 1
            for col_idx, orig_col_idx in enumerate(campaign_col_indices):
                cell_value = ws.cell(row=row_idx, column=orig_col_idx + 1).value
                sendable_ws.cell(row=sendable_count + 1, column=col_idx + 1, value=cell_value)

    # Save sendable file
    sendable_path = file_path.parent / f"{file_path.stem}_sendable{file_path.suffix}"
    sendable_wb.save(sendable_path)

    if verbose:
        print(f"\nSendable file created: {sendable_path.name} ({sendable_count} rows ready for SmartLead)")

    return {
        "sendable_file": str(sendable_path),
        "sendable_count": sendable_count,
        "columns_exported": len(campaign_headers)
    }


def process_excel_file(
    file_path: str,
    wait: bool = True,
    task_id: Optional[str] = None,
    force: bool = False,
    no_export: bool = False,
    verbose: bool = True
) -> Dict:
    """
    Process a single Excel file and add Bounce Ban verification columns.

    Args:
        file_path: Path to the Excel file
        wait: If True, wait for verification. If False, return task_id immediately.
        task_id: If provided, download results for this task_id instead of creating new task
        force: If True, allow overwriting existing verification
        no_export: If True, skip creating sendable file
        verbose: Whether to print progress

    Returns:
        Dict with processing statistics
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    if not file_path.suffix.lower() in ['.xlsx', '.xls']:
        return {"error": f"Not an Excel file: {file_path}"}

    # Create backup
    backup_path = file_path.parent / f"{file_path.stem}_backup{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    if verbose:
        print(f"Backup created: {backup_path}")

    # Load workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {"error": f"Failed to load Excel file: {str(e)}"}

    stats = {
        "file": str(file_path),
        "sheets_processed": 0,
        "rows_processed": 0,
        "verified": 0,
        "results": Counter()
    }

    # Process first sheet only
    ws = wb.active

    if ws.max_row < 2:
        return {"error": "Empty spreadsheet"}

    headers = get_headers(ws)
    if not headers or all(h is None for h in headers):
        return {"error": "No headers found"}

    # Find email column
    email_col_idx = find_column_index(headers, EMAIL_COLUMN_PATTERNS)

    if email_col_idx is None:
        return {"error": f"No email column found. Looked for: {', '.join(EMAIL_COLUMN_PATTERNS)}"}

    if verbose:
        print(f"Found email column: {headers[email_col_idx]}")

    # Find MV_Status column (REQUIRED - must run after Million Verifier)
    mv_status_col_idx = find_column_index(headers, ['mv_status', 'millionverifier_status'])

    if mv_status_col_idx is None:
        return {"error": "MV_Status column not found. Run /million-verifier first."}

    # Check for duplicate verification
    if not check_duplicate_verification(ws, headers, force):
        return {
            "error": "Bounce Ban verification already run on this file. Use --force to overwrite.",
            "duplicate_detected": True
        }

    # Find or create Bounce Ban columns
    bb_result_col_idx = find_column_index(headers, ['bb_result', 'bounceban_result'])
    bb_score_col_idx = find_column_index(headers, ['bb_score', 'bounceban_score'])
    bb_accept_all_col_idx = find_column_index(headers, ['bb_is_accept_all', 'bb_accept_all'])
    bb_role_col_idx = find_column_index(headers, ['bb_is_role', 'bb_role'])
    bb_free_col_idx = find_column_index(headers, ['bb_is_free', 'bb_free'])
    bb_disposable_col_idx = find_column_index(headers, ['bb_is_disposable', 'bb_disposable'])
    bb_sendable_col_idx = find_column_index(headers, ['bb_sendable', 'bounceban_sendable'])

    if bb_result_col_idx is None:
        bb_result_col_idx = len(headers)
        ws.cell(row=1, column=bb_result_col_idx + 1, value='BB_Result')
        headers.append('BB_Result')

    if bb_score_col_idx is None:
        bb_score_col_idx = len(headers)
        ws.cell(row=1, column=bb_score_col_idx + 1, value='BB_Score')
        headers.append('BB_Score')

    if bb_accept_all_col_idx is None:
        bb_accept_all_col_idx = len(headers)
        ws.cell(row=1, column=bb_accept_all_col_idx + 1, value='BB_Is_Accept_All')
        headers.append('BB_Is_Accept_All')

    if bb_role_col_idx is None:
        bb_role_col_idx = len(headers)
        ws.cell(row=1, column=bb_role_col_idx + 1, value='BB_Is_Role')
        headers.append('BB_Is_Role')

    if bb_free_col_idx is None:
        bb_free_col_idx = len(headers)
        ws.cell(row=1, column=bb_free_col_idx + 1, value='BB_Is_Free')
        headers.append('BB_Is_Free')

    if bb_disposable_col_idx is None:
        bb_disposable_col_idx = len(headers)
        ws.cell(row=1, column=bb_disposable_col_idx + 1, value='BB_Is_Disposable')
        headers.append('BB_Is_Disposable')

    if bb_sendable_col_idx is None:
        bb_sendable_col_idx = len(headers)
        ws.cell(row=1, column=bb_sendable_col_idx + 1, value='BB_Sendable')
        headers.append('BB_Sendable')

    # PHASE 1: Collect emails for verification (only 'ok' or 'catch_all' from MV)
    email_data = []  # List of (row_idx, email)
    for row_idx in range(2, ws.max_row + 1):
        email_value = ws.cell(row=row_idx, column=email_col_idx + 1).value
        mv_status_value = ws.cell(row=row_idx, column=mv_status_col_idx + 1).value

        if email_value and mv_status_value:
            email = str(email_value).strip()
            mv_status = str(mv_status_value).strip().lower()

            # Only verify emails marked as 'ok' or 'catch_all' by Million Verifier
            if email and '@' in email and mv_status in ['ok', 'catch_all']:
                email_data.append((row_idx, email))

    if not email_data:
        if verbose:
            print("\nNo emails to verify (all were filtered out by Million Verifier)")
            print("Bounce Ban only verifies emails with MV_Status = 'ok' or 'catch_all'")
        return {
            "success": True,
            "message": "No emails to verify (all filtered by Million Verifier)",
            "rows_processed": 0,
            "verified": 0
        }

    if verbose:
        print(f"Found {len(email_data)} emails to verify (filtered by MV_Status)")

    # PHASE 2: Verify emails via Bounce Ban
    if task_id is None:
        # Upload and verify
        emails = [email for _, email in email_data]
        result = verify_emails(emails, wait=wait, task_name=f"BB_{file_path.stem}")

        if not result['success']:
            return {"error": result.get('error', 'Verification failed')}

        if not wait:
            # Background job submitted
            return {
                "success": True,
                "task_id": result['task_id'],
                "message": f"Verification submitted (task_id: {result['task_id']}). Check status with --status {result['task_id']}"
            }

        verification_results = result['results']
        stats['results'].update(result.get('stats', {}))

    else:
        # Download existing results
        if verbose:
            print(f"Downloading results for task_id {task_id}...")

        try:
            results_list = download_results(task_id)
        except Exception as e:
            return {"error": str(e)}

        # Convert to dict keyed by email
        verification_results = {}
        for r in results_list:
            verification_results[r.email.lower()] = {
                'result': r.result,
                'score': r.score,
                'is_disposable': r.is_disposable,
                'is_accept_all': r.is_accept_all,
                'is_role': r.is_role,
                'is_free': r.is_free
            }

        # Count stats
        for r in results_list:
            stats['results'][r.result] += 1

    # PHASE 3: Write results back to spreadsheet
    mv_sendable_count = 0
    bb_sendable_count = 0

    for row_idx in range(2, ws.max_row + 1):
        email_value = ws.cell(row=row_idx, column=email_col_idx + 1).value
        mv_status_value = ws.cell(row=row_idx, column=mv_status_col_idx + 1).value

        if not email_value:
            continue

        email = str(email_value).strip()
        email_lower = email.lower()

        # Count MV sendable
        mv_sendable_col_idx = find_column_index(headers, ['mv_sendable'])
        if mv_sendable_col_idx is not None:
            mv_sendable_value = ws.cell(row=row_idx, column=mv_sendable_col_idx + 1).value
            if mv_sendable_value is True or str(mv_sendable_value).upper() == 'TRUE':
                mv_sendable_count += 1

        # Check if this email was verified by Bounce Ban
        if email_lower in verification_results:
            vr = verification_results[email_lower]
            stats["rows_processed"] += 1

            # BB_Result
            result = vr.get('result', 'unknown')
            ws.cell(row=row_idx, column=bb_result_col_idx + 1, value=result)

            # BB_Score
            score = vr.get('score', 0)
            ws.cell(row=row_idx, column=bb_score_col_idx + 1, value=score)

            # BB_Is_Accept_All
            is_accept_all = vr.get('is_accept_all', False)
            ws.cell(row=row_idx, column=bb_accept_all_col_idx + 1, value=is_accept_all)

            # BB_Is_Role
            is_role = vr.get('is_role', False)
            ws.cell(row=row_idx, column=bb_role_col_idx + 1, value=is_role)

            # BB_Is_Free
            is_free = vr.get('is_free', False)
            ws.cell(row=row_idx, column=bb_free_col_idx + 1, value=is_free)

            # BB_Is_Disposable
            is_disposable = vr.get('is_disposable', False)
            ws.cell(row=row_idx, column=bb_disposable_col_idx + 1, value=is_disposable)

            # BB_Sendable (TRUE only if result is 'deliverable')
            sendable = result == 'deliverable'
            ws.cell(row=row_idx, column=bb_sendable_col_idx + 1, value=sendable)

            if sendable:
                bb_sendable_count += 1

            stats["verified"] += 1
        else:
            # Email not verified by BB (was filtered by MV or not in verified list)
            # Leave BB columns empty for these rows
            pass

    stats["sheets_processed"] = 1
    stats["mv_sendable_count"] = mv_sendable_count
    stats["bb_sendable_count"] = bb_sendable_count

    # Save workbook
    try:
        wb.save(file_path)
        if verbose:
            print(f"Saved: {file_path}")
    except Exception as e:
        return {"error": f"Failed to save: {str(e)}", **stats}

    # PHASE 4: Export sendable file (unless --no-export)
    if not no_export and bb_sendable_count > 0:
        export_result = export_sendable_file(wb, file_path, bb_sendable_col_idx, headers, verbose)
        stats.update(export_result)
    elif no_export and verbose:
        print("\nSkipping sendable file export (--no-export enabled)")

    return stats


def print_summary(result: Dict):
    """Print summary of processing results."""
    print("\n" + "=" * 60)
    print("BOUNCE BAN - PROCESSING SUMMARY")
    print("=" * 60)

    if "error" in result:
        print(f"ERROR: {result['error']}")
        if result.get("duplicate_detected"):
            print("\nTo overwrite existing verification, use: --force")
        return

    if "message" in result:
        print(f"\n{result['message']}")
        return

    if "file" in result:
        print(f"File: {Path(result['file']).name}")
    if "sheets_processed" in result:
        print(f"Sheets processed: {result['sheets_processed']}")
    if "rows_processed" in result:
        print(f"Emails verified by BB: {result['rows_processed']}")
    if "verified" in result:
        print(f"Successfully verified: {result['verified']}")

    if "results" in result and result["results"]:
        print("\nVERIFICATION RESULTS:")
        for status, count in result["results"].items():
            print(f"  {status}: {count}")

        # Calculate sendable count
        sendable = result["results"].get('deliverable', 0)
        total = sum(result["results"].values())
        if total > 0:
            sendable_pct = (sendable / total) * 100
            print(f"\nBB SENDABLE EMAILS: {sendable} ({sendable_pct:.1f}%)")

    # Comparison with Million Verifier
    if "mv_sendable_count" in result and "bb_sendable_count" in result:
        mv_count = result["mv_sendable_count"]
        bb_count = result["bb_sendable_count"]
        print(f"\nCOMPARISON:")
        print(f"  Million Verifier marked {mv_count} as sendable")
        print(f"  Bounce Ban confirmed {bb_count} as sendable")
        if mv_count > 0:
            confirmation_pct = (bb_count / mv_count) * 100
            print(f"  Confirmation rate: {confirmation_pct:.1f}%")

    # Sendable file info
    if "sendable_file" in result:
        print(f"\nSENDABLE FILE: {Path(result['sendable_file']).name}")
        print(f"  Rows ready for SmartLead: {result.get('sendable_count', 0)}")
        print(f"  Columns exported: {result.get('columns_exported', 0)}")

    print("=" * 60)


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python add_bb_verification_column.py leads.xlsx")
        print("  python add_bb_verification_column.py leads.xlsx --no-export    # Don't create sendable file")
        print("  python add_bb_verification_column.py leads.xlsx --force        # Overwrite existing verification")
        print("  python add_bb_verification_column.py leads.xlsx --status task_id  # Check job status")
        print("  python add_bb_verification_column.py leads.xlsx --download task_id  # Download results")
        sys.exit(1)

    file_path = sys.argv[1]

    # Parse options
    wait = True
    task_id = None
    force = False
    no_export = False

    if '--no-export' in sys.argv:
        no_export = True

    if '--force' in sys.argv:
        force = True

    if '--status' in sys.argv:
        idx = sys.argv.index('--status')
        if idx + 1 < len(sys.argv):
            task_id_arg = sys.argv[idx + 1]
            status = get_task_status(task_id_arg)
            print(f"Task ID: {status.task_id}")
            print(f"Status: {status.status}")
            if status.total > 0:
                percent = (status.verified / status.total) * 100
                print(f"Progress: {percent:.1f}% ({status.verified}/{status.total})")
            print(f"Deliverable: {status.deliverable}")
            print(f"Risky: {status.risky}")
            print(f"Undeliverable: {status.undeliverable}")
            print(f"Unknown: {status.unknown}")
            sys.exit(0)

    if '--download' in sys.argv:
        idx = sys.argv.index('--download')
        if idx + 1 < len(sys.argv):
            task_id = sys.argv[idx + 1]

    # Process file
    result = process_excel_file(file_path, wait=wait, task_id=task_id, force=force, no_export=no_export)
    print_summary(result)


if __name__ == '__main__':
    main()
