"""
Add Niche Column Script
Main orchestrator that processes Excel files, scrapes/researches companies,
and adds a Verified_Niche column.
"""

import os
import sys
import time
import shutil
from pathlib import Path
from typing import Optional, List, Dict, Tuple
from collections import Counter

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import our modules
from scrape_homepage import scrape_homepage
from research_company import research_company
from categorize_company_niche import categorize_niche


# Column name patterns for finding website/domain
WEBSITE_COLUMN_PATTERNS = [
    'website', 'url', 'domain', 'web', 'site', 'homepage',
    'company_website', 'company_url', 'company_domain',
    'web_address', 'website_url'
]

# Column name patterns for finding company name
COMPANY_COLUMN_PATTERNS = [
    'company', 'company_name', 'companyname', 'name', 'organization',
    'org', 'business', 'business_name', 'firm', 'company name',
    'organization_name', 'org_name'
]

# Rate limiting
REQUEST_DELAY = 1.5  # seconds between requests


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx

    return None


def get_headers(ws) -> List[str]:
    """Get header row from worksheet."""
    return [cell.value for cell in ws[1]]


def process_excel_file(file_path: str, verbose: bool = True) -> Dict:
    """
    Process a single Excel file and add Verified_Niche column.

    Args:
        file_path: Path to the Excel file
        verbose: Whether to print progress

    Returns:
        Dict with processing statistics
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    if not file_path.suffix.lower() in ['.xlsx', '.xls']:
        return {"error": f"Not an Excel file: {file_path}"}

    # Create backup
    backup_path = file_path.parent / f"{file_path.stem}_backup{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    if verbose:
        print(f"Backup created: {backup_path}")

    # Load workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {"error": f"Failed to load Excel file: {str(e)}"}

    stats = {
        "file": str(file_path),
        "sheets_processed": 0,
        "rows_processed": 0,
        "successful_categorizations": 0,
        "failed_categorizations": 0,
        "niche_distribution": Counter()
    }

    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.max_row < 2:  # Empty or header only
            continue

        headers = get_headers(ws)
        if not headers or all(h is None for h in headers):
            continue

        # Find website column
        website_col_idx = find_column_index(headers, WEBSITE_COLUMN_PATTERNS)

        # Find company name column
        company_col_idx = find_column_index(headers, COMPANY_COLUMN_PATTERNS)

        if website_col_idx is None and company_col_idx is None:
            if verbose:
                print(f"  Sheet '{sheet_name}': No website or company column found, skipping")
            continue

        # Find or create Verified_Niche column
        niche_col_idx = find_column_index(headers, ['verified_niche', 'verified niche'])

        if niche_col_idx is None:
            # Add new column
            niche_col_idx = len(headers)
            ws.cell(row=1, column=niche_col_idx + 1, value='Verified_Niche')

        if verbose:
            print(f"  Processing sheet '{sheet_name}' ({ws.max_row - 1} rows)...")

        stats["sheets_processed"] += 1

        # Process each row
        for row_idx in range(2, ws.max_row + 1):
            # Get website URL
            website = None
            if website_col_idx is not None:
                cell_value = ws.cell(row=row_idx, column=website_col_idx + 1).value
                if cell_value:
                    website = str(cell_value).strip()

            # Get company name
            company_name = None
            if company_col_idx is not None:
                cell_value = ws.cell(row=row_idx, column=company_col_idx + 1).value
                if cell_value:
                    company_name = str(cell_value).strip()

            # Skip if no data
            if not website and not company_name:
                ws.cell(row=row_idx, column=niche_col_idx + 1, value='Insufficient Data')
                stats["failed_categorizations"] += 1
                continue

            stats["rows_processed"] += 1

            if verbose and stats["rows_processed"] % 10 == 0:
                print(f"    Processed {stats['rows_processed']} rows...")

            # Get content for analysis
            content = None

            if website:
                # Try scraping the website
                scraped = scrape_homepage(website)
                if scraped.success:
                    content = scraped.to_text()

            if not content and company_name:
                # Fall back to research by company name
                research = research_company(company_name)
                if research.success:
                    content = research.to_text()
                    # If research found a website, we could scrape it
                    if research.inferred_website and not website:
                        scraped = scrape_homepage(research.inferred_website)
                        if scraped.success:
                            content = scraped.to_text()

            # Categorize
            if content:
                result = categorize_niche(content, company_name or '')

                if result.success:
                    ws.cell(row=row_idx, column=niche_col_idx + 1, value=result.niche)
                    stats["successful_categorizations"] += 1
                    stats["niche_distribution"][result.niche] += 1
                else:
                    ws.cell(row=row_idx, column=niche_col_idx + 1, value='Categorization Failed')
                    stats["failed_categorizations"] += 1
            else:
                ws.cell(row=row_idx, column=niche_col_idx + 1, value='Research Required')
                stats["failed_categorizations"] += 1

            # Rate limiting
            time.sleep(REQUEST_DELAY)

    # Save workbook
    try:
        wb.save(file_path)
        if verbose:
            print(f"Saved: {file_path}")
    except Exception as e:
        return {"error": f"Failed to save: {str(e)}", **stats}

    return stats


def process_folder(folder_path: str, verbose: bool = True) -> List[Dict]:
    """
    Process all Excel files in a folder recursively.

    Args:
        folder_path: Path to the folder
        verbose: Whether to print progress

    Returns:
        List of stats for each file processed
    """
    folder_path = Path(folder_path)
    if not folder_path.exists():
        return [{"error": f"Folder not found: {folder_path}"}]

    if not folder_path.is_dir():
        return [{"error": f"Not a folder: {folder_path}"}]

    # Find all Excel files
    excel_files = list(folder_path.glob('**/*.xlsx')) + list(folder_path.glob('**/*.xls'))

    # Filter out backup files
    excel_files = [f for f in excel_files if '_backup' not in f.stem]

    if not excel_files:
        return [{"error": f"No Excel files found in: {folder_path}"}]

    if verbose:
        print(f"Found {len(excel_files)} Excel files to process")

    results = []
    for file_path in excel_files:
        if verbose:
            print(f"\nProcessing: {file_path.name}")
        result = process_excel_file(str(file_path), verbose=verbose)
        results.append(result)

    return results


def print_summary(results: List[Dict]):
    """Print summary of processing results."""
    print("\n" + "=" * 60)
    print("PROCESSING SUMMARY")
    print("=" * 60)

    total_rows = 0
    total_success = 0
    total_failed = 0
    all_niches = Counter()

    for result in results:
        if "error" in result and result.get("rows_processed", 0) == 0:
            print(f"ERROR: {result['error']}")
        else:
            if "file" in result:
                print(f"\nFile: {Path(result['file']).name}")
            if "sheets_processed" in result:
                print(f"  Sheets processed: {result['sheets_processed']}")
            if "rows_processed" in result:
                print(f"  Rows processed: {result['rows_processed']}")
                total_rows += result['rows_processed']
            if "successful_categorizations" in result:
                print(f"  Successful: {result['successful_categorizations']}")
                total_success += result['successful_categorizations']
            if "failed_categorizations" in result:
                print(f"  Failed: {result['failed_categorizations']}")
                total_failed += result['failed_categorizations']
            if "niche_distribution" in result:
                all_niches.update(result['niche_distribution'])

    print("\n" + "-" * 60)
    print("TOTALS")
    print(f"  Total rows processed: {total_rows}")
    print(f"  Total successful: {total_success}")
    print(f"  Total failed: {total_failed}")

    if all_niches:
        print("\nNICHE DISTRIBUTION (Top 20):")
        for niche, count in all_niches.most_common(20):
            print(f"  {niche}: {count}")

    print("=" * 60)


def main():
    if len(sys.argv) < 2:
        print("Usage: python add_niche_column.py <file_or_folder_path>")
        print("\nExamples:")
        print("  python add_niche_column.py companies.xlsx")
        print("  python add_niche_column.py ./leads_folder/")
        sys.exit(1)

    path = sys.argv[1]

    if Path(path).is_file():
        result = process_excel_file(path)
        print_summary([result])
    elif Path(path).is_dir():
        results = process_folder(path)
        print_summary(results)
    else:
        print(f"Error: Path not found: {path}")
        sys.exit(1)


if __name__ == '__main__':
    main()
