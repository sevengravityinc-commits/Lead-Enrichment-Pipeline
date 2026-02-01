"""
Identify Decision-Makers Script
Fast batch AI identification of decision-makers using job titles and company context.
Uses OpenRouter API with batch processing for cost-effective classification.
"""

import os
import sys
import json
import time
import csv
import argparse
import requests
from pathlib import Path
from typing import List, Dict, Optional
from collections import Counter
from dotenv import load_dotenv

load_dotenv()

try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Configuration
DEFAULT_BATCH_SIZE = 20  # Smaller batch size for more reliable results
DEFAULT_MODEL = "openai/gpt-4o-mini"
API_DELAY = 0.5  # Seconds between API calls
MAX_RETRIES = 3  # Number of retry attempts for invalid responses

OPENROUTER_API_KEY = os.getenv('OPENROUTER_API_KEY')


# Decision-Maker Classification Prompt
DECISION_MAKER_PROMPT = """Classify contacts as decision-makers for B2B sales.

DECISION-MAKERS (answer "Yes"):
• C-Suite: CEO, CFO, CMO, CTO, COO, CIO
• Founders/Owners
• VPs/Vice Presidents (all functions)
• Directors (exclude junior/associate directors)
• Heads: Head of Marketing, Head of Sales, Head of Growth

NOT DECISION-MAKERS (answer "No"):
• Managers (unless "Senior Manager" + strategic function)
• Coordinators, Specialists, Analysts
• Junior/Assistant/Associate titles
• Individual contributors

CONFIDENCE LEVELS:
• "High" = Clear C-Suite/Founder/VP title
• "Medium" = Director/Head/Senior Manager
• "Low" = Ambiguous or missing title

OUTPUT FORMAT - Return a JSON array with these exact fields:
{
  "decision_maker": "Yes" or "No" (ONLY these two values),
  "confidence": "High" or "Medium" or "Low" (ONLY these three values)
}

CRITICAL:
- Return EXACTLY one result per input contact
- Use "Yes"/"No" for decision_maker (NOT "True", "Medium", "Low", etc.)
- Use "High"/"Medium"/"Low" for confidence
- Return ONLY the JSON array, no other text

Example (CEO → Yes/High, Coordinator → No/High, Director → Yes/Medium):
[
  {"decision_maker": "Yes", "confidence": "High"},
  {"decision_maker": "No", "confidence": "High"},
  {"decision_maker": "Yes", "confidence": "Medium"}
]"""


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns (case-insensitive, partial matching)."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx
    return None


def format_contact_for_classification(contact: Dict) -> str:
    """Format contact data for AI prompt with all available context."""
    parts = [f"Title: {contact.get('title', 'N/A')}"]

    if contact.get('company'):
        parts.append(f"Company: {contact['company']}")
    if contact.get('industry'):
        parts.append(f"Industry: {contact['industry']}")
    if contact.get('name'):
        parts.append(f"Name: {contact['name']}")

    return " | ".join(parts)


def batch_classify_decision_makers(contacts: List[Dict], model: str) -> List[Dict]:
    """
    Send a batch of contacts to OpenRouter and get decision-maker classifications.

    Args:
        contacts: List of dicts with 'title' and optional context keys
        model: AI model to use

    Returns:
        List of {"decision_maker": "Yes/No", "confidence": "High/Medium/Low"}
    """
    # Format contacts for prompt
    contact_lines = [format_contact_for_classification(c) for c in contacts]
    companies_json = json.dumps(contact_lines)

    # Retry loop (up to MAX_RETRIES attempts)
    for attempt in range(MAX_RETRIES):
        try:
            response = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": model,
                    "max_tokens": len(contacts) * 30,
                    "messages": [
                        {"role": "system", "content": DECISION_MAKER_PROMPT},
                        {"role": "user", "content": f"Classify these contacts:\n{companies_json}"}
                    ]
                },
                timeout=60
            )

            response.raise_for_status()
            data = response.json()
            response_text = data['choices'][0]['message']['content'].strip()

            # Parse JSON (handle markdown code blocks)
            if response_text.startswith("```"):
                response_text = response_text.split("```")[1]
                if response_text.startswith("json"):
                    response_text = response_text[4:]
                response_text = response_text.strip()

            results = json.loads(response_text)

            # Validate results
            valid_dm = {"Yes", "No"}
            valid_conf = {"High", "Medium", "Low"}
            all_valid = True

            for r in results:
                if r.get('decision_maker') not in valid_dm or r.get('confidence') not in valid_conf:
                    all_valid = False
                    break

            if all_valid and len(results) == len(contacts):
                return results
            else:
                if attempt < MAX_RETRIES - 1:
                    print(f"  Warning: Invalid response format, retrying (attempt {attempt + 2}/{MAX_RETRIES})...")
                    print(f"  Debug: Got {len(results)} results for {len(contacts)} contacts")
                    print(f"  Debug: all_valid={all_valid}")
                    if not all_valid and len(results) > 0:
                        print(f"  Debug: First invalid result: {results[0] if len(results) > 0 else 'N/A'}")
                    time.sleep(1)
                    continue
                else:
                    # All retries failed - mark as failed
                    print(f"  Error: Failed to get valid classification after {MAX_RETRIES} attempts")
                    print(f"  Debug: Got {len(results)} results for {len(contacts)} contacts")
                    print(f"  Debug: all_valid={all_valid}")
                    return [{"decision_maker": "No", "confidence": "Low"} for _ in contacts]

        except json.JSONDecodeError as e:
            if attempt < MAX_RETRIES - 1:
                print(f"  Warning: JSON parse error, retrying (attempt {attempt + 2}/{MAX_RETRIES})...")
                time.sleep(1)
                continue
            else:
                print(f"  Error: JSON parse failed after {MAX_RETRIES} attempts: {e}")
                return [{"decision_maker": "No", "confidence": "Low"} for _ in contacts]

        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                print(f"  Warning: API error, retrying (attempt {attempt + 2}/{MAX_RETRIES})...")
                time.sleep(1)
                continue
            else:
                print(f"  Error: API failed after {MAX_RETRIES} attempts: {e}")
                return [{"decision_maker": "No", "confidence": "Low"} for _ in contacts]

    # Should never reach here, but safety fallback
    return [{"decision_maker": "No", "confidence": "Low"} for _ in contacts]


def create_decision_makers_csv(file_path: Path, ws, headers: List[str], dm_col_idx: int) -> tuple:
    """
    Create CSV file with only decision-maker rows (Decision_Maker = "Yes").

    Returns:
        (csv_file_path, decision_maker_count)
    """
    base_name = file_path.stem
    csv_filename = f"{base_name}_DECISION_MAKERS.csv"
    csv_path = file_path.parent / csv_filename

    # Collect all Yes rows
    decision_maker_rows = []

    for row_idx in range(2, ws.max_row + 1):
        dm_value = ws.cell(row=row_idx, column=dm_col_idx + 1).value

        if dm_value == "Yes":
            row_data = [ws.cell(row=row_idx, column=col).value
                       for col in range(1, len(headers) + 1)]
            decision_maker_rows.append(row_data)

    # Write CSV
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(decision_maker_rows)

    print(f"\nDecision makers CSV created: {csv_filename} ({len(decision_maker_rows)} rows)")

    return csv_path, len(decision_maker_rows)


def process_excel_file(
    file_path: str,
    batch_size: int = DEFAULT_BATCH_SIZE,
    model: str = DEFAULT_MODEL,
    verbose: bool = True
) -> Dict:
    """
    Process Excel file and add Decision_Maker + Confidence columns.

    Args:
        file_path: Path to Excel file
        batch_size: Contacts per API call
        model: AI model to use
        verbose: Whether to print progress

    Returns:
        Dict with processing statistics
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    if not OPENROUTER_API_KEY:
        return {"error": "OPENROUTER_API_KEY not set in environment"}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    # Create backup
    backup_path = file_path.parent / f"{file_path.stem}_backup{file_path.suffix}"
    if not backup_path.exists():
        import shutil
        shutil.copy2(file_path, backup_path)
        if verbose:
            print(f"Backup created: {backup_path.name}")

    # Load workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {"error": f"Failed to load Excel file: {e}"}

    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Intelligent column detection
    columns = {
        'title': find_column_index(headers, ['title', 'job_title', 'job title', 'jobtitle', 'position', 'role', 'job_position', 'designation']),
        'company': find_column_index(headers, ['company', 'company_name', 'company name', 'organization', 'account']),
        'name': find_column_index(headers, ['name', 'contact_name', 'contact name', 'full_name', 'person', 'contact']),
        'industry': find_column_index(headers, ['industry', 'sector', 'vertical', 'company_industry']),
    }

    # Title is required
    if columns['title'] is None:
        wb.close()
        return {"error": "Title column not found. This column is required for decision-maker identification. Please ensure your file has a column named 'Title', 'Job Title', or similar."}

    # Report detected columns
    if verbose:
        print(f"\nDetected columns:")
        print(f"  Title: column {columns['title'] + 1} (required)")
        for key in ['company', 'industry', 'name']:
            if columns[key] is not None:
                print(f"  {key.replace('_', ' ').title()}: column {columns[key] + 1} (optional context)")

    # Find or create Decision_Maker column
    dm_col = find_column_index(headers, ['decision_maker', 'decision maker'])
    if dm_col is None:
        dm_col = len(headers)
        ws.cell(row=1, column=dm_col + 1, value='Decision_Maker')
        headers.append('Decision_Maker')

    # Find or create Confidence column
    conf_col = find_column_index(headers, ['confidence', 'confidence level'])
    if conf_col is None:
        conf_col = len(headers)
        ws.cell(row=1, column=conf_col + 1, value='Confidence')
        headers.append('Confidence')

    # Get total rows
    total_rows = ws.max_row - 1

    if verbose:
        print(f"\nProcessing {total_rows} rows in batches of {batch_size}...")
        print(f"Estimated API calls: {(total_rows + batch_size - 1) // batch_size}")
        print(f"Model: {model}")
        print()

    # Statistics
    stats = {
        "total_rows": total_rows,
        "processed": 0,
        "skipped": 0,
        "batches": 0,
        "distribution": Counter(),
        "confidence_distribution": Counter()
    }

    # Process in batches
    batch = []
    batch_rows = []

    for row_idx in range(2, ws.max_row + 1):
        # Skip if already classified with valid values
        existing_dm = ws.cell(row=row_idx, column=dm_col + 1).value
        existing_conf = ws.cell(row=row_idx, column=conf_col + 1).value

        if existing_dm in ['Yes', 'No'] and existing_conf in ['High', 'Medium', 'Low']:
            stats['skipped'] += 1
            stats['distribution'][existing_dm] += 1
            stats['confidence_distribution'][existing_conf] += 1
            continue

        # Get contact data
        title = ws.cell(row=row_idx, column=columns['title'] + 1).value

        # Handle missing title
        if not title or str(title).strip() == "":
            ws.cell(row=row_idx, column=dm_col + 1, value='No')
            ws.cell(row=row_idx, column=conf_col + 1, value='Low')
            stats['processed'] += 1
            stats['distribution']['No'] += 1
            stats['confidence_distribution']['Low'] += 1
            continue

        # Build contact dict with all available context
        contact_data = {'title': str(title).strip()}

        for key in ['company', 'industry', 'name']:
            if columns[key] is not None:
                value = ws.cell(row=row_idx, column=columns[key] + 1).value
                if value:
                    contact_data[key] = str(value).strip()

        batch.append(contact_data)
        batch_rows.append(row_idx)

        # Process batch when full
        if len(batch) >= batch_size:
            stats['batches'] += 1

            if verbose:
                progress = (stats['processed'] + stats['skipped']) / total_rows * 100
                print(f"  Batch {stats['batches']}: rows {batch_rows[0]}-{batch_rows[-1]} ({progress:.1f}% complete)")

            # Get classifications from AI
            results = batch_classify_decision_makers(batch, model)

            # Write results
            for row, result in zip(batch_rows, results):
                ws.cell(row=row, column=dm_col + 1, value=result['decision_maker'])
                ws.cell(row=row, column=conf_col + 1, value=result['confidence'])
                stats['distribution'][result['decision_maker']] += 1
                stats['confidence_distribution'][result['confidence']] += 1
                stats['processed'] += 1

            # Save progress
            try:
                wb.save(file_path)
            except Exception as e:
                print(f"  Warning: Could not save progress: {e}")

            # Clear batch
            batch = []
            batch_rows = []

            # Rate limiting
            time.sleep(API_DELAY)

    # Process remaining batch
    if batch:
        stats['batches'] += 1

        if verbose:
            print(f"  Batch {stats['batches']} (final): rows {batch_rows[0]}-{batch_rows[-1]}")

        results = batch_classify_decision_makers(batch, model)

        for row, result in zip(batch_rows, results):
            ws.cell(row=row, column=dm_col + 1, value=result['decision_maker'])
            ws.cell(row=row, column=conf_col + 1, value=result['confidence'])
            stats['distribution'][result['decision_maker']] += 1
            stats['confidence_distribution'][result['confidence']] += 1
            stats['processed'] += 1

    # Final save
    try:
        wb.save(file_path)
        if verbose:
            print(f"\nSaved: {file_path}")
    except Exception as e:
        wb.close()
        return {"error": f"Failed to save: {e}", **stats}

    # Create decision makers CSV
    csv_path, dm_count = create_decision_makers_csv(file_path, ws, headers, dm_col)

    wb.close()

    stats['csv_file'] = str(csv_path)
    stats['decision_maker_count'] = dm_count

    return stats


def main():
    parser = argparse.ArgumentParser(
        description='Identify decision-makers in lead lists using AI classification'
    )
    parser.add_argument('file', help='Path to Excel file')
    parser.add_argument(
        '--batch-size',
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help=f'Contacts per API call (default: {DEFAULT_BATCH_SIZE})'
    )
    parser.add_argument(
        '--model',
        default=DEFAULT_MODEL,
        help=f'AI model to use (default: {DEFAULT_MODEL})'
    )

    args = parser.parse_args()

    print(f"\n{'='*60}")
    print("DECISION-MAKER IDENTIFICATION")
    print(f"{'='*60}")
    print(f"File: {args.file}")
    print(f"Batch size: {args.batch_size}")
    print(f"Model: {args.model}")

    start_time = time.time()

    result = process_excel_file(
        args.file,
        batch_size=args.batch_size,
        model=args.model,
        verbose=True
    )

    elapsed = time.time() - start_time

    print(f"\n{'='*60}")
    print("RESULTS")
    print(f"{'='*60}")

    if "error" in result:
        print(f"Error: {result['error']}")
        sys.exit(1)

    print(f"Total rows: {result['total_rows']}")
    print(f"Processed: {result['processed']}")
    print(f"Skipped (already done): {result['skipped']}")
    print(f"API batches: {result['batches']}")
    print(f"Time elapsed: {elapsed:.1f} seconds")

    # Decision-maker distribution
    print(f"\nDecision-Maker Distribution:")
    for category in ['Yes', 'No']:
        count = result['distribution'].get(category, 0)
        pct = count / result['total_rows'] * 100 if result['total_rows'] > 0 else 0
        print(f"  {category}: {count} ({pct:.1f}%)")

    # Confidence distribution
    print(f"\nConfidence Distribution:")
    for conf in ['High', 'Medium', 'Low']:
        count = result['confidence_distribution'].get(conf, 0)
        pct = count / result['total_rows'] * 100 if result['total_rows'] > 0 else 0
        print(f"  {conf}: {count} ({pct:.1f}%)")

    print(f"\nDecision makers CSV saved to: {result['csv_file']}")


if __name__ == '__main__':
    main()
