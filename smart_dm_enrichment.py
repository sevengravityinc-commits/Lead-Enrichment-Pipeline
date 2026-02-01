"""
Smart Decision Maker Enrichment Script
Processes lead lists to classify, verify, and enrich decision makers.

Features:
- Process up to 3 files at once
- AI-based decision maker classification
- BlitzAPI verification and lookup
- Fuzzy name matching
- Domain lookup via Google search
- Credit estimation before processing
"""

import os
import sys
import csv
import json
import time
import argparse
import requests
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field
from dotenv import load_dotenv

load_dotenv()

# Optional imports
try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from blitz_api import BlitzAPI, BlitzAPIError
    BLITZ_AVAILABLE = True
except ImportError:
    BLITZ_AVAILABLE = False

try:
    from rapidfuzz import fuzz
    FUZZY_AVAILABLE = True
except ImportError:
    FUZZY_AVAILABLE = False


# Configuration
OPENROUTER_API_KEY = os.getenv('OPENROUTER_API_KEY')
AI_MODEL = "openai/gpt-4o-mini"
BATCH_SIZE = 20
API_DELAY = 0.5
NAME_MATCH_THRESHOLD = 80

# Decision Maker Classification Prompt
DECISION_MAKER_PROMPT = """Classify contacts as decision-makers for B2B sales.

DECISION-MAKERS (answer "Yes"):
• C-Suite: CEO, CFO, CMO, CTO, COO, CIO
• Founders/Owners
• VPs/Vice Presidents (all functions)
• Directors (exclude junior/associate directors)
• Heads: Head of Marketing, Head of Sales, Head of Growth

NOT DECISION-MAKERS (answer "No"):
• Managers (unless "Senior Manager" + strategic function)
• Coordinators, Specialists, Analysts
• Junior/Assistant/Associate titles
• Individual contributors

INDUSTRY-SPECIFIC CONSIDERATION:
If industry is provided, consider industry-specific decision maker titles:
• Healthcare: Chief Medical Officer, VP Clinical, Medical Director, Director of Clinical Operations
• Technology/SaaS: VP Engineering, VP Product, Chief Architect, Engineering Director
• Manufacturing: VP Operations, VP Supply Chain, Plant Director, Operations Director
• Retail/E-commerce: VP Merchandising, Chief Merchandising Officer, Retail Director
• Finance/Banking: Chief Risk Officer, VP Compliance, Chief Credit Officer
• Media/Advertising: VP Creative, Chief Creative Officer, Media Director

Example: "Director of Clinical Operations" in Healthcare = likely DM (High confidence)
Same title in a Tech company = likely NOT DM (Low confidence)

CONFIDENCE LEVELS:
• "High" = Clear C-Suite/Founder/VP title
• "Medium" = Director/Head/Senior Manager or industry-specific DM title
• "Low" = Ambiguous or missing title

OUTPUT FORMAT - Return a JSON array with these exact fields:
{
  "decision_maker": "Yes" or "No" (ONLY these two values),
  "confidence": "High" or "Medium" or "Low" (ONLY these three values)
}

CRITICAL:
- Return EXACTLY one result per input contact
- Use "Yes"/"No" for decision_maker (NOT "True", "Medium", "Low", etc.)
- Use "High"/"Medium"/"Low" for confidence
- Return ONLY the JSON array, no other text
"""


@dataclass
class ProcessingStats:
    """Track processing statistics"""
    total_contacts: int = 0
    skipped: int = 0  # Already processed rows (skip logic)
    verified: int = 0
    email_updated: int = 0
    email_added: int = 0
    not_found: int = 0
    not_dm: int = 0
    new_dms_added: int = 0
    no_domain: int = 0
    errors: int = 0
    credits_used: int = 0


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns (case-insensitive)."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]
    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx
    return None


def clean_domain(url: str) -> str:
    """Extract clean domain from URL or domain string."""
    if not url:
        return ""
    url = str(url).strip().lower()
    for prefix in ['https://', 'http://', 'www.']:
        if url.startswith(prefix):
            url = url[len(prefix):]
    url = url.split('/')[0]
    url = url.split(':')[0]
    return url


def find_domain_from_company_name(company_name: str) -> Optional[str]:
    """
    Find company domain via Google search.
    Returns domain or None if not found.
    """
    if not company_name:
        return None

    try:
        # Use a simple approach - search for company website
        search_query = f"{company_name} official website"

        # Use requests to search (simplified approach)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }

        # Try DuckDuckGo instant answer API (free, no rate limits)
        response = requests.get(
            f"https://api.duckduckgo.com/?q={company_name}&format=json",
            headers=headers,
            timeout=5
        )

        if response.status_code == 200:
            data = response.json()
            # Check for official URL in results
            if data.get('Results'):
                for result in data['Results']:
                    url = result.get('FirstURL', '')
                    if url:
                        domain = clean_domain(url)
                        if domain and '.' in domain:
                            return domain

            # Check related topics
            if data.get('RelatedTopics'):
                for topic in data['RelatedTopics'][:3]:
                    url = topic.get('FirstURL', '')
                    if url and company_name.lower().replace(' ', '') in url.lower():
                        domain = clean_domain(url)
                        if domain and '.' in domain:
                            return domain

        return None

    except Exception as e:
        print(f"    Domain lookup error for {company_name}: {e}")
        return None


def classify_contacts_batch(contacts: List[Dict]) -> List[Dict]:
    """
    Send a batch of contacts to AI for decision-maker classification.
    Returns list of {"decision_maker": "Yes/No", "confidence": "High/Medium/Low"}
    """
    if not OPENROUTER_API_KEY:
        return [{"decision_maker": "No", "confidence": "Low"} for _ in contacts]

    # Format contacts for classification
    contact_lines = []
    for c in contacts:
        title = c.get('title', '') or ''
        company = c.get('company', '') or ''
        name = c.get('name', '') or ''
        industry = c.get('industry', '') or ''
        line = f"Title: {title}"
        if company:
            line += f" | Company: {company}"
        if name:
            line += f" | Name: {name}"
        if industry:
            line += f" | Industry: {industry}"
        contact_lines.append(line)

    contacts_json = json.dumps(contact_lines)

    for attempt in range(3):
        try:
            response = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": AI_MODEL,
                    "max_tokens": len(contacts) * 30,
                    "messages": [
                        {"role": "system", "content": DECISION_MAKER_PROMPT},
                        {"role": "user", "content": f"Classify these contacts:\n{contacts_json}"}
                    ]
                },
                timeout=60
            )

            response.raise_for_status()
            data = response.json()
            response_text = data['choices'][0]['message']['content'].strip()

            # Parse JSON (handle markdown code blocks)
            if response_text.startswith("```"):
                response_text = response_text.split("```")[1]
                if response_text.startswith("json"):
                    response_text = response_text[4:]
                response_text = response_text.strip()

            results = json.loads(response_text)

            # Validate results
            if len(results) == len(contacts):
                return results

        except Exception as e:
            if attempt < 2:
                time.sleep(1)
                continue

    # Fallback: return Low confidence for all
    return [{"decision_maker": "No", "confidence": "Low"} for _ in contacts]


def fuzzy_name_match(name: str, dm_list: List[Dict], threshold: int = NAME_MATCH_THRESHOLD) -> Optional[Dict]:
    """
    Find best fuzzy match for a name in the DM list.
    Returns the matched DM dict or None.
    """
    if not name or not dm_list:
        return None

    name = str(name).strip().lower()

    best_match = None
    best_score = 0

    for dm in dm_list:
        dm_name = dm.get('full_name', '') or ''
        dm_name = str(dm_name).strip().lower()

        if not dm_name:
            continue

        if FUZZY_AVAILABLE:
            # Use rapidfuzz for better matching
            score = fuzz.ratio(name, dm_name)
            # Also try partial matching for cases like "John" vs "John Smith"
            partial_score = fuzz.partial_ratio(name, dm_name)
            score = max(score, partial_score)
        else:
            # Simple matching fallback
            if name == dm_name:
                score = 100
            elif name in dm_name or dm_name in name:
                score = 80
            else:
                # Check first/last name
                name_parts = name.split()
                dm_parts = dm_name.split()
                matching = sum(1 for p in name_parts if any(p in dp for dp in dm_parts))
                score = (matching / max(len(name_parts), 1)) * 100

        if score > best_score:
            best_score = score
            best_match = dm

    if best_score >= threshold:
        return best_match

    return None


def estimate_credits(file_paths: List[str], verbose: bool = True) -> Dict:
    """
    Estimate credits needed for processing.
    Returns dict with estimates and warnings.
    """
    total_contacts = 0
    unique_domains = set()

    for file_path in file_paths:
        file_path = Path(file_path)
        if not file_path.exists():
            continue

        try:
            if file_path.suffix.lower() == '.csv':
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                headers = rows[0] if rows else []
                data_rows = rows[1:] if len(rows) > 1 else []
            else:
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                data_rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
                wb.close()

            total_contacts += len(data_rows)

            # Find domain column
            domain_col = find_column_index(headers, ['website', 'domain', 'url', 'web'])
            company_col = find_column_index(headers, ['company', 'company_name', 'organization'])

            for row in data_rows:
                domain = None
                if domain_col is not None and domain_col < len(row):
                    domain = clean_domain(row[domain_col])
                if domain:
                    unique_domains.add(domain)
                elif company_col is not None and company_col < len(row):
                    # Will need domain lookup
                    unique_domains.add(f"lookup_{row[company_col]}")

        except Exception as e:
            if verbose:
                print(f"Error reading {file_path}: {e}")

    # Estimate credits:
    # - Domain lookup: 1 credit per unique domain (for companies without domain)
    # - DM search: 1 credit per search + 1 credit per person found
    # - Email enrichment: 1 credit per person

    domain_lookups = len([d for d in unique_domains if d.startswith('lookup_')])
    company_searches = len(unique_domains)

    # Worst case: 2 DMs per company with emails
    min_credits = company_searches * 3  # 1 search + 2 people minimum
    max_credits = company_searches * 7 + domain_lookups  # search + 4 people + emails + domain lookups

    estimate = {
        'total_contacts': total_contacts,
        'unique_companies': len(unique_domains),
        'domain_lookups_needed': domain_lookups,
        'min_credits': min_credits,
        'max_credits': max_credits,
        'avg_credits': (min_credits + max_credits) // 2
    }

    return estimate


def process_file(
    file_path: str,
    company_cache: Dict,
    max_new_per_company: int = 2,
    batch_size: int = 20,
    verbose: bool = True
) -> Tuple[List[Dict], ProcessingStats]:
    """
    Process a single file for DM enrichment.

    Args:
        file_path: Path to Excel/CSV file
        batch_size: Number of contacts per AI API call (default 20, use 10 for large files)
        company_cache: Shared cache of company DM lookups (domain -> results)
        max_new_per_company: Max new DMs to add per company
        verbose: Print progress

    Returns:
        Tuple of (enriched rows, stats)
    """
    file_path = Path(file_path)
    stats = ProcessingStats()

    if not file_path.exists():
        return [], stats

    # Load file
    if file_path.suffix.lower() == '.csv':
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
    else:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data_rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        wb.close()

    # Detect columns
    cols = {
        'title': find_column_index(headers, ['title', 'job_title', 'job title', 'position', 'role']),
        'email': find_column_index(headers, ['email', 'email_address', 'work_email', 'e-mail']),
        'company': find_column_index(headers, ['company', 'company_name', 'organization', 'account']),
        'domain': find_column_index(headers, ['website', 'domain', 'url', 'web', 'company_website']),
        'first_name': find_column_index(headers, ['first_name', 'firstname', 'first name', 'fname']),
        'last_name': find_column_index(headers, ['last_name', 'lastname', 'last name', 'lname']),
        'name': find_column_index(headers, ['name', 'full_name', 'fullname', 'contact_name', 'contact']),
        'industry': find_column_index(headers, ['industry', 'sector', 'vertical', 'company_industry']),
        'decision_status': find_column_index(headers, ['decision_status']),  # For skip logic
    }

    if verbose:
        print(f"\nDetected columns: {[k for k, v in cols.items() if v is not None]}")

    # Initialize BlitzAPI
    api = None
    if BLITZ_AVAILABLE:
        try:
            api = BlitzAPI()
        except Exception as e:
            if verbose:
                print(f"BlitzAPI init error: {e}")

    # Process contacts
    contacts_for_classification = []
    contact_data = []

    for row in data_rows:
        contact = {
            'original_row': list(row),
            'title': row[cols['title']] if cols['title'] is not None and cols['title'] < len(row) else None,
            'email': row[cols['email']] if cols['email'] is not None and cols['email'] < len(row) else None,
            'company': row[cols['company']] if cols['company'] is not None and cols['company'] < len(row) else None,
            'domain': clean_domain(row[cols['domain']]) if cols['domain'] is not None and cols['domain'] < len(row) else None,
            'name': None,
            'industry': row[cols['industry']] if cols['industry'] is not None and cols['industry'] < len(row) else None,
            'existing_status': row[cols['decision_status']] if cols['decision_status'] is not None and cols['decision_status'] < len(row) else None,
        }

        # Build full name
        if cols['name'] is not None and cols['name'] < len(row):
            contact['name'] = row[cols['name']]
        elif cols['first_name'] is not None or cols['last_name'] is not None:
            first = row[cols['first_name']] if cols['first_name'] is not None and cols['first_name'] < len(row) else ''
            last = row[cols['last_name']] if cols['last_name'] is not None and cols['last_name'] < len(row) else ''
            contact['name'] = f"{first or ''} {last or ''}".strip()

        contact_data.append(contact)
        contacts_for_classification.append({
            'title': contact['title'],
            'company': contact['company'],
            'name': contact['name'],
            'industry': contact['industry']
        })

    stats.total_contacts = len(contact_data)

    if verbose:
        print(f"Total contacts: {stats.total_contacts}")

    # Batch classify all contacts
    if verbose:
        print("Classifying contacts as decision makers...")

    classifications = []
    for i in range(0, len(contacts_for_classification), batch_size):
        batch = contacts_for_classification[i:i+batch_size]
        batch_results = classify_contacts_batch(batch)
        classifications.extend(batch_results)
        if verbose:
            print(f"  Classified {min(i+batch_size, len(contacts_for_classification))}/{len(contacts_for_classification)}")

    # Process each contact based on classification
    enriched_rows = []
    new_rows_to_add = []

    for idx, (contact, classification) in enumerate(zip(contact_data, classifications)):
        # Skip logic: if already processed, keep original row and skip
        if contact.get('existing_status') and str(contact['existing_status']).strip():
            output_row = list(contact['original_row'])
            output_row.append(contact['existing_status'])  # Keep existing decision_status
            output_row.append('')  # Empty confidence for skipped rows
            enriched_rows.append(output_row)
            stats.skipped += 1
            if verbose and (idx + 1) % 50 == 0:
                print(f"  Skipped {stats.skipped} already-processed rows")
            continue

        is_dm = classification.get('decision_maker') == 'Yes'
        confidence = classification.get('confidence', 'Low')

        # Treat Low confidence as NOT a DM
        if confidence == 'Low':
            is_dm = False

        # Get domain (try lookup if missing)
        domain = contact['domain']
        if not domain and contact['company']:
            if verbose:
                print(f"  [{idx+1}] Looking up domain for {contact['company']}...", end=" ")
            domain = find_domain_from_company_name(contact['company'])
            if domain:
                contact['domain'] = domain
                if verbose:
                    print(f"Found: {domain}")
            else:
                if verbose:
                    print("Not found")

        # Prepare the output row
        output_row = list(contact['original_row'])

        # Handle based on classification
        if not is_dm:
            # NOT a decision maker
            decision_status = 'not_dm'
            stats.not_dm += 1

            # Find new DMs if we have domain
            if domain and api:
                new_dms = lookup_company_dms(api, domain, company_cache, max_new_per_company, verbose)
                for dm in new_dms:
                    new_rows_to_add.append(create_new_dm_row(headers, cols, contact, dm))
                    stats.new_dms_added += 1
            elif not domain:
                decision_status = 'no_domain'
                stats.no_domain += 1

        else:
            # IS a decision maker - verify via BlitzAPI
            if domain and api:
                # Look up DMs at company
                dms_at_company = lookup_company_dms(api, domain, company_cache, 5, verbose=False)

                # Try to match by name
                matched_dm = fuzzy_name_match(contact['name'], dms_at_company)

                if matched_dm:
                    # Found a match
                    blitz_email = matched_dm.get('email', '')
                    original_email = contact['email'] or ''

                    if not original_email and blitz_email:
                        # Email was missing, add it
                        decision_status = 'email_added'
                        if cols['email'] is not None:
                            output_row[cols['email']] = blitz_email
                        stats.email_added += 1

                    elif blitz_email and original_email.lower() != blitz_email.lower():
                        # Email changed
                        decision_status = 'email_updated'
                        if cols['email'] is not None:
                            output_row[cols['email']] = blitz_email
                        stats.email_updated += 1

                    else:
                        # Email verified
                        decision_status = 'verified'
                        stats.verified += 1

                    # Add 1 more DM
                    new_dms = [dm for dm in dms_at_company if dm != matched_dm][:1]
                    for dm in new_dms:
                        new_rows_to_add.append(create_new_dm_row(headers, cols, contact, dm))
                        stats.new_dms_added += 1

                else:
                    # Not found at company
                    decision_status = 'not_found'
                    stats.not_found += 1

                    # Add 2 new DMs
                    for dm in dms_at_company[:max_new_per_company]:
                        new_rows_to_add.append(create_new_dm_row(headers, cols, contact, dm))
                        stats.new_dms_added += 1
            else:
                # No domain or API
                if not domain:
                    decision_status = 'no_domain'
                    stats.no_domain += 1
                else:
                    decision_status = 'verified'  # Can't verify without API
                    stats.verified += 1

        # Add decision_status and confidence to row
        output_row.append(decision_status)
        output_row.append(confidence)
        enriched_rows.append(output_row)

        if verbose and (idx + 1) % 10 == 0:
            print(f"  Processed {idx+1}/{len(contact_data)} contacts")

        time.sleep(API_DELAY)

    # Add new rows
    for new_row in new_rows_to_add:
        enriched_rows.append(new_row)

    # Add decision_status and dm_confidence headers
    headers_out = list(headers) + ['decision_status', 'dm_confidence']

    return headers_out, enriched_rows, stats


def lookup_company_dms(api, domain: str, cache: Dict, max_results: int, verbose: bool = False) -> List[Dict]:
    """Look up decision makers at a company, using cache."""
    if domain in cache:
        return cache[domain][:max_results]

    try:
        if verbose:
            print(f"    Looking up DMs at {domain}...", end=" ")

        results = api.search_decision_makers(
            company_domain=domain,
            with_email=True
        )

        cache[domain] = results or []

        if verbose:
            print(f"Found {len(results or [])} DMs")

        return (results or [])[:max_results]

    except BlitzAPIError as e:
        if verbose:
            print(f"Error: {e.message}")
        if e.status_code == 402:
            raise  # Out of credits - propagate
        cache[domain] = []
        return []
    except Exception as e:
        if verbose:
            print(f"Error: {e}")
        cache[domain] = []
        return []


def create_new_dm_row(headers: List, cols: Dict, original_contact: Dict, dm: Dict) -> List:
    """Create a new row for a found decision maker."""
    # Start with empty row
    new_row = [None] * len(headers)

    # Fill in company info from original
    if cols['company'] is not None:
        new_row[cols['company']] = original_contact['company']
    if cols['domain'] is not None:
        new_row[cols['domain']] = original_contact['domain'] or dm.get('company_domain', '')

    # Fill in DM info
    if cols['email'] is not None:
        new_row[cols['email']] = dm.get('email', '')
    if cols['title'] is not None:
        new_row[cols['title']] = dm.get('title', '')
    if cols['name'] is not None:
        new_row[cols['name']] = dm.get('full_name', '')
    if cols['first_name'] is not None:
        new_row[cols['first_name']] = dm.get('first_name', '')
    if cols['last_name'] is not None:
        new_row[cols['last_name']] = dm.get('last_name', '')

    # Add decision_status and dm_confidence
    new_row.append('new')
    new_row.append('High')  # BlitzAPI-found DMs are high confidence

    return new_row


def save_results(file_path: Path, headers: List, rows: List) -> str:
    """Save enriched results to Excel file."""
    output_path = file_path.parent / f"{file_path.stem}_ENRICHED.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched"

    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=value)

    wb.save(output_path)
    return str(output_path)


def create_dm_csv(file_path: Path, headers: List, rows: List) -> str:
    """
    Create CSV with only decision makers (verified, email_updated, email_added, new).

    Args:
        file_path: Original input file path
        headers: Column headers
        rows: All enriched rows

    Returns:
        Path to created CSV file
    """
    dm_statuses = {'verified', 'email_updated', 'email_added', 'new'}

    # Find decision_status column index (second to last)
    status_idx = len(headers) - 2  # decision_status is second to last, dm_confidence is last

    # Filter to only DM rows
    dm_rows = [r for r in rows if len(r) > status_idx and r[status_idx] in dm_statuses]

    if not dm_rows:
        return None

    # Create CSV
    csv_path = file_path.parent / f"{file_path.stem}_DECISION_MAKERS.csv"

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(dm_rows)

    return str(csv_path)


def process_multiple_files(
    file_paths: List[str],
    max_new_per_company: int = 2,
    batch_size: int = 20,
    export_dm_csv: bool = False,
    verbose: bool = True
) -> Dict:
    """
    Process multiple files for DM enrichment.

    Args:
        file_paths: List of paths to Excel/CSV files
        max_new_per_company: Max new DMs to add per company
        batch_size: Contacts per AI API call (default 20, use 10 for large files)
        export_dm_csv: Also create DM-only CSV file
        verbose: Print progress

    Returns:
        Dict with results and stats
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    if not BLITZ_AVAILABLE:
        return {"error": "BlitzAPI not available. Check blitz_api.py exists."}

    # Check credits
    try:
        api = BlitzAPI()
        credits_info = api.get_key_info()
        available_credits = credits_info.remaining_credits

        estimate = estimate_credits(file_paths, verbose=False)

        if verbose:
            print(f"\n{'='*60}")
            print("CREDIT ESTIMATION")
            print(f"{'='*60}")
            print(f"Total contacts: {estimate['total_contacts']}")
            print(f"Unique companies: {estimate['unique_companies']}")
            print(f"Estimated credits: {estimate['min_credits']} - {estimate['max_credits']}")
            print(f"Available credits: {available_credits}")

            if available_credits < estimate['min_credits']:
                print(f"\n WARNING: May not have enough credits!")
                print(f"         Need at least {estimate['min_credits']}, have {available_credits}")
            print(f"{'='*60}\n")

    except Exception as e:
        if verbose:
            print(f"Could not check credits: {e}")

    # Shared cache for company lookups
    company_cache = {}

    # Process each file
    results = {
        'files_processed': [],
        'total_stats': ProcessingStats(),
        'output_files': []
    }

    for file_path in file_paths:
        file_path = Path(file_path)
        if not file_path.exists():
            if verbose:
                print(f"File not found: {file_path}")
            continue

        if verbose:
            print(f"\n{'='*60}")
            print(f"PROCESSING: {file_path.name}")
            print(f"{'='*60}")

        try:
            headers, rows, stats = process_file(
                str(file_path),
                company_cache,
                max_new_per_company,
                batch_size,
                verbose
            )

            # Save results
            output_path = save_results(file_path, headers, rows)
            results['output_files'].append(output_path)
            results['files_processed'].append(str(file_path))

            # Optionally create DM-only CSV
            if export_dm_csv:
                dm_csv_path = create_dm_csv(file_path, headers, rows)
                if dm_csv_path:
                    results.setdefault('dm_csv_files', []).append(dm_csv_path)
                    if verbose:
                        print(f"DM CSV created: {dm_csv_path}")

            # Aggregate stats
            results['total_stats'].total_contacts += stats.total_contacts
            results['total_stats'].skipped += stats.skipped
            results['total_stats'].verified += stats.verified
            results['total_stats'].email_updated += stats.email_updated
            results['total_stats'].email_added += stats.email_added
            results['total_stats'].not_found += stats.not_found
            results['total_stats'].not_dm += stats.not_dm
            results['total_stats'].new_dms_added += stats.new_dms_added
            results['total_stats'].no_domain += stats.no_domain
            results['total_stats'].errors += stats.errors

            if verbose:
                print(f"\nFile stats:")
                print(f"  Verified: {stats.verified}")
                print(f"  Email updated: {stats.email_updated}")
                print(f"  Email added: {stats.email_added}")
                print(f"  Not found: {stats.not_found}")
                print(f"  Not DM: {stats.not_dm}")
                print(f"  New DMs added: {stats.new_dms_added}")
                print(f"  No domain: {stats.no_domain}")
                print(f"Output: {output_path}")

        except BlitzAPIError as e:
            if e.status_code == 402:
                print(f"\n Out of credits! Stopping.")
                results['error'] = "Out of credits"
                break
            else:
                if verbose:
                    print(f"Error processing {file_path}: {e.message}")
                results['total_stats'].errors += 1

        except Exception as e:
            if verbose:
                print(f"Error processing {file_path}: {e}")
            results['total_stats'].errors += 1

    # Get credits used
    try:
        credits_after = api.get_key_info().remaining_credits
        results['total_stats'].credits_used = available_credits - credits_after
    except:
        pass

    return results


def main():
    parser = argparse.ArgumentParser(
        description='Smart Decision Maker Enrichment - Process lead lists'
    )
    parser.add_argument('files', nargs='+', help='Excel/CSV files to process (up to 3)')
    parser.add_argument(
        '--max-new',
        type=int,
        default=2,
        help='Max new DMs to add per company (default: 2)'
    )
    parser.add_argument(
        '--batch-size',
        type=int,
        default=20,
        help='Contacts per AI API call (default: 20, use 10 for large files)'
    )
    parser.add_argument(
        '--export-dm-csv',
        action='store_true',
        help='Also create DM-only CSV file'
    )
    parser.add_argument(
        '--estimate-only',
        action='store_true',
        help='Only show credit estimate, do not process'
    )

    args = parser.parse_args()

    if len(args.files) > 3:
        print("Error: Maximum 3 files allowed")
        sys.exit(1)

    print(f"\n{'='*60}")
    print("SMART DECISION MAKER ENRICHMENT")
    print(f"{'='*60}")
    print(f"Files: {', '.join(args.files)}")
    print(f"Max new DMs per company: {args.max_new}")
    print(f"Batch size: {args.batch_size}")
    print(f"Export DM CSV: {args.export_dm_csv}")

    if args.estimate_only:
        estimate = estimate_credits(args.files)
        print(f"\nEstimate:")
        print(f"  Total contacts: {estimate['total_contacts']}")
        print(f"  Unique companies: {estimate['unique_companies']}")
        print(f"  Domain lookups needed: {estimate['domain_lookups_needed']}")
        print(f"  Credits needed: {estimate['min_credits']} - {estimate['max_credits']}")
        sys.exit(0)

    start_time = time.time()

    result = process_multiple_files(
        args.files,
        max_new_per_company=args.max_new,
        batch_size=args.batch_size,
        export_dm_csv=args.export_dm_csv,
        verbose=True
    )

    elapsed = time.time() - start_time

    print(f"\n{'='*60}")
    print("FINAL RESULTS")
    print(f"{'='*60}")

    if "error" in result:
        print(f"Error: {result['error']}")

    stats = result['total_stats']
    print(f"Files processed: {len(result['files_processed'])}")
    print(f"Total contacts: {stats.total_contacts}")
    print(f"  Skipped (already processed): {stats.skipped}")
    print(f"  Verified: {stats.verified}")
    print(f"  Email updated: {stats.email_updated}")
    print(f"  Email added: {stats.email_added}")
    print(f"  Not found at company: {stats.not_found}")
    print(f"  Not decision makers: {stats.not_dm}")
    print(f"  No domain: {stats.no_domain}")
    print(f"New DMs added: {stats.new_dms_added}")
    print(f"Credits used: {stats.credits_used}")
    print(f"Time elapsed: {elapsed:.1f} seconds")

    if result.get('output_files'):
        print(f"\nOutput files:")
        for f in result['output_files']:
            print(f"  {f}")

    if result.get('dm_csv_files'):
        print(f"\nDM-only CSV files:")
        for f in result['dm_csv_files']:
            print(f"  {f}")


if __name__ == '__main__':
    main()
