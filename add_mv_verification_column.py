"""
Add Million Verifier Columns Script
Processes Excel files and adds email verification status columns using Million Verifier API.
"""

import os
import sys
import shutil
import time
from pathlib import Path
from typing import Optional, List, Dict
from collections import Counter

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from millionverifier_api import verify_emails, get_file_status, download_results


# Column name patterns for finding email
EMAIL_COLUMN_PATTERNS = [
    'email', 'e-mail', 'email address', 'work email', 'business email',
    'contact email', 'primary email', 'mail'
]


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx

    return None


def get_headers(ws) -> List[str]:
    """Get header row from worksheet."""
    return [cell.value for cell in ws[1]]


def check_duplicate_verification(ws, headers: List[str], force: bool = False) -> bool:
    """
    Check if Million Verifier verification has already been run on this file.

    Args:
        ws: Worksheet to check
        headers: Header row
        force: If True, allow overwrite

    Returns:
        True if safe to proceed, False if duplicate detected and not forced
    """
    mv_status_col_idx = find_column_index(headers, ['mv_status', 'millionverifier_status'])

    if mv_status_col_idx is None:
        return True  # No MV columns yet, safe to proceed

    # Check if column has data (more than just header)
    has_data = False
    for row_idx in range(2, min(ws.max_row + 1, 12)):  # Check first 10 rows
        cell_value = ws.cell(row=row_idx, column=mv_status_col_idx + 1).value
        if cell_value and str(cell_value).strip():
            has_data = True
            break

    if not has_data:
        return True  # Column exists but empty, safe to proceed

    if force:
        print("WARNING: Overwriting existing Million Verifier verification (--force enabled)")
        time.sleep(3)  # 3-second warning
        return True

    return False  # Duplicate detected, not forced


def process_excel_file(file_path: str, wait: bool = True, file_id: Optional[int] = None, force: bool = False, verbose: bool = True) -> Dict:
    """
    Process a single Excel file and add email verification columns.

    Args:
        file_path: Path to the Excel file
        wait: If True, wait for verification. If False, return file_id immediately.
        file_id: If provided, download results for this file_id instead of uploading
        verbose: Whether to print progress

    Returns:
        Dict with processing statistics
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    if not file_path.suffix.lower() in ['.xlsx', '.xls']:
        return {"error": f"Not an Excel file: {file_path}"}

    # Create backup
    backup_path = file_path.parent / f"{file_path.stem}_backup{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    if verbose:
        print(f"Backup created: {backup_path}")

    # Load workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {"error": f"Failed to load Excel file: {str(e)}"}

    stats = {
        "file": str(file_path),
        "sheets_processed": 0,
        "rows_processed": 0,
        "verified": 0,
        "results": Counter()
    }

    # Process first sheet only (typically all data is in first sheet)
    ws = wb.active

    if ws.max_row < 2:
        return {"error": "Empty spreadsheet"}

    headers = get_headers(ws)
    if not headers or all(h is None for h in headers):
        return {"error": "No headers found"}

    # Find email column
    email_col_idx = find_column_index(headers, EMAIL_COLUMN_PATTERNS)

    if email_col_idx is None:
        return {"error": f"No email column found. Looked for: {', '.join(EMAIL_COLUMN_PATTERNS)}"}

    if verbose:
        print(f"Found email column: {headers[email_col_idx]}")

    # Check for duplicate verification
    if not check_duplicate_verification(ws, headers, force):
        return {
            "error": "Million Verifier verification already run on this file. Use --force to overwrite.",
            "duplicate_detected": True
        }

    # Find or create verification columns
    mv_status_col_idx = find_column_index(headers, ['mv_status', 'millionverifier_status'])
    mv_quality_col_idx = find_column_index(headers, ['mv_quality', 'millionverifier_quality'])
    mv_sendable_col_idx = find_column_index(headers, ['mv_sendable', 'millionverifier_sendable'])

    if mv_status_col_idx is None:
        mv_status_col_idx = len(headers)
        ws.cell(row=1, column=mv_status_col_idx + 1, value='MV_Status')
        headers.append('MV_Status')

    if mv_quality_col_idx is None:
        mv_quality_col_idx = len(headers)
        ws.cell(row=1, column=mv_quality_col_idx + 1, value='MV_Quality')
        headers.append('MV_Quality')

    if mv_sendable_col_idx is None:
        mv_sendable_col_idx = len(headers)
        ws.cell(row=1, column=mv_sendable_col_idx + 1, value='MV_Sendable')
        headers.append('MV_Sendable')

    # PHASE 1: Collect emails and their row positions
    email_data = []  # List of (row_idx, email)
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=email_col_idx + 1).value
        if cell_value:
            email = str(cell_value).strip()
            if email and '@' in email:
                email_data.append((row_idx, email))

    if not email_data:
        return {"error": "No valid emails found"}

    if verbose:
        print(f"Found {len(email_data)} emails to verify")

    # Determine if we should wait based on list size
    should_wait = wait and len(email_data) < 5000

    # PHASE 2: Verify emails via Million Verifier
    if file_id is None:
        # Upload and verify
        emails = [email for _, email in email_data]
        result = verify_emails(emails, wait=should_wait)

        if not result['success']:
            return {"error": result.get('error', 'Verification failed')}

        if not should_wait:
            # Background job submitted
            return {
                "success": True,
                "file_id": result['file_id'],
                "message": f"Verification submitted (file_id: {result['file_id']}). Check status with --status {result['file_id']}"
            }

        verification_results = result['results']
        stats.update(result.get('stats', {}))

    else:
        # Download existing results
        if verbose:
            print(f"Downloading results for file_id {file_id}...")

        try:
            results_list = download_results(file_id)
        except Exception as e:
            return {"error": str(e)}

        # Convert to dict keyed by email
        verification_results = {}
        for r in results_list:
            verification_results[r.email.lower()] = {
                'quality': r.quality,
                'result': r.result,
                'free': r.free,
                'role': r.role
            }

        # Count stats
        for r in results_list:
            stats['results'][r.result] += 1

    # PHASE 3: Write results back to spreadsheet
    for row_idx, email in email_data:
        stats["rows_processed"] += 1

        email_lower = email.lower()
        if email_lower in verification_results:
            vr = verification_results[email_lower]

            # MV_Status
            status = vr.get('result', 'unknown')
            ws.cell(row=row_idx, column=mv_status_col_idx + 1, value=status)

            # MV_Quality
            quality = vr.get('quality', 'unknown')
            ws.cell(row=row_idx, column=mv_quality_col_idx + 1, value=quality)

            # MV_Sendable (TRUE only if status is 'ok')
            sendable = status == 'ok'
            ws.cell(row=row_idx, column=mv_sendable_col_idx + 1, value=sendable)

            stats["verified"] += 1
        else:
            # Email not found in results
            ws.cell(row=row_idx, column=mv_status_col_idx + 1, value='error')
            ws.cell(row=row_idx, column=mv_quality_col_idx + 1, value='unknown')
            ws.cell(row=row_idx, column=mv_sendable_col_idx + 1, value=False)

    stats["sheets_processed"] = 1

    # Save workbook
    try:
        wb.save(file_path)
        if verbose:
            print(f"Saved: {file_path}")
    except Exception as e:
        return {"error": f"Failed to save: {str(e)}", **stats}

    return stats


def print_summary(result: Dict):
    """Print summary of processing results."""
    print("\n" + "=" * 60)
    print("MILLION VERIFIER - PROCESSING SUMMARY")
    print("=" * 60)

    if "error" in result:
        print(f"ERROR: {result['error']}")
        if result.get("duplicate_detected"):
            print("\nTo overwrite existing verification, use: --force")
        return

    if "message" in result:
        print(f"\n{result['message']}")
        return

    if "file" in result:
        print(f"File: {Path(result['file']).name}")
    if "sheets_processed" in result:
        print(f"Sheets processed: {result['sheets_processed']}")
    if "rows_processed" in result:
        print(f"Emails processed: {result['rows_processed']}")
    if "verified" in result:
        print(f"Successfully verified: {result['verified']}")

    if "results" in result and result["results"]:
        print("\nVERIFICATION RESULTS:")
        for status, count in result["results"].items():
            print(f"  {status}: {count}")

        # Calculate sendable count
        sendable = result["results"].get('ok', 0)
        total = sum(result["results"].values())
        if total > 0:
            sendable_pct = (sendable / total) * 100
            print(f"\nSENDABLE EMAILS: {sendable} ({sendable_pct:.1f}%)")

    print("=" * 60)


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python add_mv_verification_column.py leads.xlsx")
        print("  python add_mv_verification_column.py leads.xlsx --no-wait    # Submit and return file_id")
        print("  python add_mv_verification_column.py leads.xlsx --force      # Overwrite existing verification")
        print("  python add_mv_verification_column.py leads.xlsx --status 12345  # Check job status")
        print("  python add_mv_verification_column.py leads.xlsx --download 12345  # Download results")
        sys.exit(1)

    file_path = sys.argv[1]

    # Parse options
    wait = True
    file_id = None
    force = False

    if '--no-wait' in sys.argv:
        wait = False

    if '--force' in sys.argv:
        force = True

    if '--status' in sys.argv:
        idx = sys.argv.index('--status')
        if idx + 1 < len(sys.argv):
            file_id_arg = int(sys.argv[idx + 1])
            status = get_file_status(file_id_arg)
            print(f"File ID: {status.file_id}")
            print(f"Status: {status.status}")
            print(f"Progress: {status.percent:.1f}% ({status.verified}/{status.total})")
            if status.error:
                print(f"Error: {status.error}")
            sys.exit(0)

    if '--download' in sys.argv:
        idx = sys.argv.index('--download')
        if idx + 1 < len(sys.argv):
            file_id = int(sys.argv[idx + 1])

    # Process file
    result = process_excel_file(file_path, wait=wait, file_id=file_id, force=force)
    print_summary(result)


if __name__ == '__main__':
    main()
