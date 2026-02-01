"""
Generate Email Sequences Script - Google Drive Version
Creates 3-email sequences (A/B/C variants) for Seven Gravity cold outreach.
Exports to Google Docs (email sequences) and Google Sheets (lead list) organized by niche.
"""

import os
import sys
import json
from pathlib import Path
from typing import Optional, List, Dict
from collections import defaultdict, Counter
from datetime import datetime

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from google_drive_helper import get_or_create_campaign_folders, move_file_to_folder
from google_docs_helper import create_document, update_document
from google_sheets_helper import create_spreadsheet, add_sheet, write_to_sheet, format_header_row, delete_default_sheet


# Column name patterns
EMAIL_COLUMN_PATTERNS = [
    'email', 'e-mail', 'email address', 'work email', 'business email'
]
FIRST_NAME_PATTERNS = [
    'first name', 'firstname', 'first_name', 'fname', 'given name'
]
LAST_NAME_PATTERNS = [
    'last name', 'lastname', 'last_name', 'lname', 'surname'
]
COMPANY_COLUMN_PATTERNS = [
    'company', 'company_name', 'companyname', 'name', 'organization'
]
CLEAN_COMPANY_PATTERNS = [
    'clean_company_name', 'clean company name', 'clean_company'
]
NICHE_PATTERNS = [
    'verified_niche', 'niche', 'industry', 'category', 'vertical'
]


def load_sequences(json_path: str = None) -> Dict:
    """Load cold email sequences from JSON."""
    if json_path is None:
        json_path = Path(__file__).parent / "cold_email_sequences.json"

    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx

    return None


def get_headers(ws) -> List[str]:
    """Get header row from worksheet."""
    return [cell.value for cell in ws[1]]


def match_sequence_to_niche(niche: str, sequences_data: Dict) -> Dict:
    """
    Match a niche to the best sequence template.

    Args:
        niche: The verified niche from categorization
        sequences_data: Loaded sequences JSON

    Returns:
        Matched sequence or fallback generic template
    """
    niche_lower = niche.lower().strip() if niche else ""

    # Try exact match first
    matched_sequences = []
    for sequence in sequences_data['sequences']:
        for industry in sequence['industry']:
            if niche_lower == industry.lower():
                matched_sequences.append(sequence)
                break

    # Try partial match if no exact match
    if not matched_sequences:
        for sequence in sequences_data['sequences']:
            for industry in sequence['industry']:
                if niche_lower in industry.lower() or industry.lower() in niche_lower:
                    matched_sequences.append(sequence)
                    break

    # Return first matched sequence or fallback to generic
    if matched_sequences:
        return matched_sequences[0]
    else:
        # Find generic fallback
        for sequence in sequences_data['sequences']:
            if 'Generic' in sequence['industry']:
                return sequence
        # Ultimate fallback
        return sequences_data['sequences'][0]


def replace_niche_in_template(template_text: str, niche: str) -> str:
    """
    Replace {niche} placeholder with actual niche name.

    Args:
        template_text: Text containing {niche} placeholders
        niche: Actual niche name (e.g., "Marketing")

    Returns:
        Text with {niche} replaced
    """
    return template_text.replace('{niche}', niche)


def format_email_sequences_doc(niche: str, sequence: Dict) -> str:
    """
    Format email sequences for Google Doc.

    Args:
        niche: Niche name (e.g., "Marketing")
        sequence: Sequence template from JSON

    Returns:
        Formatted text for Google Doc with {niche} replaced by actual niche name
    """
    date = datetime.now().strftime('%Y-%m-%d')

    lines = [
        f"EMAIL SEQUENCES - {niche}",
        f"Generated: {date}",
        "",
    ]

    variant_names = {
        'a': 'SEQUENCE A (Problem-focused, direct)',
        'b': 'SEQUENCE B (Context-add, example-heavy)',
        'c': 'SEQUENCE C (Short, punchy, curiosity-driven)'
    }

    for variant_letter in ['a', 'b', 'c']:
        lines.append("═" * 60)
        lines.append(variant_names[variant_letter])
        lines.append("═" * 60)
        lines.append("")

        for email_num in range(1, 4):
            email_key = f'email{email_num}'
            variant_key = f'variant_{variant_letter}'

            email_data = sequence[email_key][variant_key]

            # Email number and type
            if email_num == 1:
                email_type = "INITIAL OUTREACH"
            elif email_num == 2:
                email_type = "FOLLOW-UP (Day 3-5)"
            else:
                email_type = "BREAKUP (Day 7-10)"

            lines.append(f"EMAIL {email_num} - {email_type}")

            # Subject (with {niche} replaced)
            subject = replace_niche_in_template(email_data['subject'], niche)
            lines.append(f"Subject: {subject}")

            # Body (with {niche} replaced)
            body = replace_niche_in_template(email_data['body'], niche)
            lines.append("Body:")
            lines.append(body)
            lines.append("")

            if email_num < 3:
                lines.append("---")
                lines.append("")

        lines.append("")

    return '\n'.join(lines)


def process_excel_file(file_path: str, verbose: bool = True) -> Dict:
    """
    Process Excel file and generate email sequences in Google Drive.

    Args:
        file_path: Path to the Excel file
        verbose: Whether to print progress

    Returns:
        Dict with processing statistics and folder URLs
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    if not file_path.suffix.lower() in ['.xlsx', '.xls']:
        return {"error": f"Not an Excel file: {file_path}"}

    # Load sequences
    try:
        sequences_data = load_sequences()
        if verbose:
            print(f"Loaded {len(sequences_data['sequences'])} Seven Gravity sequences")
    except Exception as e:
        return {"error": f"Failed to load sequences: {str(e)}"}

    # Load workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {"error": f"Failed to load Excel file: {str(e)}"}

    ws = wb.active

    if ws.max_row < 2:
        return {"error": "Empty spreadsheet"}

    headers = get_headers(ws)
    if not headers or all(h is None for h in headers):
        return {"error": "No headers found"}

    # Find required columns
    email_col_idx = find_column_index(headers, EMAIL_COLUMN_PATTERNS)
    first_name_col_idx = find_column_index(headers, FIRST_NAME_PATTERNS)
    last_name_col_idx = find_column_index(headers, LAST_NAME_PATTERNS)
    niche_col_idx = find_column_index(headers, NICHE_PATTERNS)

    if email_col_idx is None:
        return {"error": f"No email column found. Looked for: {', '.join(EMAIL_COLUMN_PATTERNS)}"}

    if niche_col_idx is None:
        return {"error": f"No niche column found. Run /lead-niche-categorizer first."}

    # Find optional columns
    company_col_idx = find_column_index(headers, COMPANY_COLUMN_PATTERNS)
    clean_company_col_idx = find_column_index(headers, CLEAN_COMPANY_PATTERNS)

    # Prefer clean company name
    company_source_idx = clean_company_col_idx if clean_company_col_idx is not None else company_col_idx

    if verbose:
        print(f"Found email column: {headers[email_col_idx]}")
        if first_name_col_idx is not None:
            print(f"Found first name column: {headers[first_name_col_idx]}")
        if last_name_col_idx is not None:
            print(f"Found last name column: {headers[last_name_col_idx]}")
        if company_source_idx is not None:
            print(f"Found company column: {headers[company_source_idx]}")
        print(f"Found niche column: {headers[niche_col_idx]}")

    # Group leads by niche
    leads_by_niche = defaultdict(list)

    for row_idx in range(2, ws.max_row + 1):
        email = ws.cell(row=row_idx, column=email_col_idx + 1).value
        if not email or not str(email).strip():
            continue

        niche = ws.cell(row=row_idx, column=niche_col_idx + 1).value
        niche = str(niche).strip() if niche else ""

        # Skip if no niche
        if not niche or niche.lower() in ['insufficient data', 'categorization failed', 'unknown']:
            continue

        # Get other fields
        first_name = ws.cell(row=row_idx, column=first_name_col_idx + 1).value if first_name_col_idx else None
        last_name = ws.cell(row=row_idx, column=last_name_col_idx + 1).value if last_name_col_idx else None
        company = ws.cell(row=row_idx, column=company_source_idx + 1).value if company_source_idx else None
        clean_company = ws.cell(row=row_idx, column=clean_company_col_idx + 1).value if clean_company_col_idx else None

        lead = {
            'email': str(email).strip(),
            'first_name': str(first_name).strip() if first_name else "",
            'last_name': str(last_name).strip() if last_name else "",
            'company': str(company).strip() if company else "",
            'clean_company_name': str(clean_company).strip() if clean_company else "",
            'niche': niche
        }

        leads_by_niche[niche].append(lead)

    if not leads_by_niche:
        return {"error": "No leads with valid niches found"}

    if verbose:
        print(f"\nGrouped leads into {len(leads_by_niche)} niches:")
        for niche, leads in leads_by_niche.items():
            print(f"  {niche}: {len(leads)} leads")

    # Process each niche
    stats = {
        "file": str(file_path),
        "niches_processed": 0,
        "total_leads": 0,
        "sequences_used": Counter(),
        "campaigns": []
    }

    for niche, leads in leads_by_niche.items():
        if verbose:
            print(f"\n{'='*60}")
            print(f"Processing {niche} niche ({len(leads)} leads)...")

        try:
            # Step 1: Create folder structure
            folder_result = get_or_create_campaign_folders(niche)

            # Step 2: Match sequence and generate doc content
            sequence = match_sequence_to_niche(niche, sequences_data)
            stats["sequences_used"][sequence['id']] += len(leads)

            # Step 3: Create Google Doc with email sequences
            doc_title = f"Email Sequences - {niche}"
            doc_result = create_document(doc_title)
            doc_content = format_email_sequences_doc(niche, sequence)
            update_document(doc_result['document_id'], doc_content)

            # Move doc to folder
            move_file_to_folder(doc_result['document_id'], folder_result['date_folder_id'])

            if verbose:
                print(f"  ✓ Created email sequences doc: {doc_result['url']}")

            # Step 4: Create Google Sheet with lead list
            sheet_title = f"Lead List - {niche}"
            sheet_result = create_spreadsheet(sheet_title)

            # Prepare data for sheet
            sheet_headers = ['Email', 'First Name', 'Last Name', 'Company', 'Clean_Company_Name', 'Niche']
            sheet_data = [sheet_headers]

            for lead in leads:
                sheet_data.append([
                    lead['email'],
                    lead['first_name'],
                    lead['last_name'],
                    lead['company'],
                    lead['clean_company_name'],
                    lead['niche']
                ])

            # Write to sheet
            write_to_sheet(sheet_result['spreadsheet_id'], 'Sheet1', sheet_data)

            # Format header row
            # Get sheet ID for Sheet1 (default is 0)
            format_header_row(sheet_result['spreadsheet_id'], 0)

            # Move sheet to folder
            move_file_to_folder(sheet_result['spreadsheet_id'], folder_result['date_folder_id'])

            if verbose:
                print(f"  ✓ Created lead list sheet: {sheet_result['url']}")

            # Add to campaign results
            stats["campaigns"].append({
                "niche": niche,
                "leads_count": len(leads),
                "folder_url": folder_result['date_folder_url'],
                "doc_url": doc_result['url'],
                "sheet_url": sheet_result['url'],
                "sequence_id": sequence['id']
            })

            stats["niches_processed"] += 1
            stats["total_leads"] += len(leads)

        except Exception as e:
            if verbose:
                print(f"  ✗ Error processing {niche}: {str(e)}")
            stats["campaigns"].append({
                "niche": niche,
                "error": str(e)
            })

    return stats


def print_summary(result: Dict):
    """Print summary of processing results."""
    print("\n" + "=" * 60)
    print("SEVEN GRAVITY EMAIL SEQUENCES - GOOGLE DRIVE")
    print("=" * 60)

    if "error" in result:
        print(f"ERROR: {result['error']}")
        return

    if "niches_processed" in result:
        print(f"Niches processed: {result['niches_processed']}")
    if "total_leads" in result:
        print(f"Total leads: {result['total_leads']}")

    if "sequences_used" in result and result["sequences_used"]:
        print("\nSEQUENCES USED:")
        for seq_id, count in result["sequences_used"].most_common():
            print(f"  {seq_id}: {count} leads")

    if "campaigns" in result and result["campaigns"]:
        print("\nCAMPAIGNS CREATED:")
        for campaign in result["campaigns"]:
            if "error" in campaign:
                print(f"\n  {campaign['niche']}: ERROR - {campaign['error']}")
            else:
                print(f"\n  {campaign['niche']} ({campaign['leads_count']} leads):")
                print(f"    Folder: {campaign['folder_url']}")
                print(f"    Email Sequences: {campaign['doc_url']}")
                print(f"    Lead List: {campaign['sheet_url']}")

    print("\nSEQUENCE STRUCTURE:")
    print("  3 sequences per niche (A/B/C):")
    print("    - Sequence A: Email 1, Email 2, Email 3 (Problem-focused)")
    print("    - Sequence B: Email 1, Email 2, Email 3 (Context-add)")
    print("    - Sequence C: Email 1, Email 2, Email 3 (Curiosity-driven)")
    print("  Total: 9 email templates per niche")

    print("\nKEY FEATURES:")
    print("  ✓ Niche hardcoded in email copy (not a placeholder)")
    print("  ✓ {first_name} and {company_name} remain as placeholders")
    print("  ✓ Organized by: Email Campaigns/{Niche}/{Date}/")
    print("  ✓ Ready for spintax addition and SmartLead upload")

    print("=" * 60)


def generate_sequences_for_niche(niche_name: str, verbose: bool = True) -> Dict:
    """
    Generate email sequence templates for a given niche (without lead file).

    Args:
        niche_name: Name of the niche (e.g., "Marketing", "SaaS")
        verbose: Whether to print progress

    Returns:
        Dict with doc URL and folder URL
    """
    if verbose:
        print(f"Generating email sequences for {niche_name} niche...")

    # Load sequences
    try:
        sequences_data = load_sequences()
        if verbose:
            print(f"Loaded {len(sequences_data['sequences'])} Seven Gravity sequences")
    except Exception as e:
        return {"error": f"Failed to load sequences: {str(e)}"}

    try:
        # Step 1: Match sequence template
        sequence = match_sequence_to_niche(niche_name, sequences_data)
        if verbose:
            print(f"Matched to sequence: {sequence['id']}")

        # Step 2: Create folder structure
        folder_result = get_or_create_campaign_folders(niche_name)
        if verbose:
            print(f"Created folder: {folder_result['date_folder_url']}")

        # Step 3: Create Google Doc with email sequences
        doc_title = f"Email Sequences - {niche_name}"
        doc_result = create_document(doc_title)
        doc_content = format_email_sequences_doc(niche_name, sequence)
        update_document(doc_result['document_id'], doc_content)

        # Move doc to folder
        move_file_to_folder(doc_result['document_id'], folder_result['date_folder_id'])

        if verbose:
            print(f"Created email sequences doc: {doc_result['url']}")

        return {
            "niche": niche_name,
            "folder_url": folder_result['date_folder_url'],
            "doc_url": doc_result['url'],
            "sequence_id": sequence['id']
        }

    except Exception as e:
        return {"error": f"Failed to generate sequences: {str(e)}"}


def print_niche_summary(result: Dict):
    """Print summary for niche-only generation."""
    print("\n" + "=" * 60)
    print("SEVEN GRAVITY EMAIL SEQUENCES - NICHE TEMPLATES")
    print("=" * 60)

    if "error" in result:
        print(f"ERROR: {result['error']}")
        return

    print(f"Niche: {result['niche']}")
    print(f"Sequence: {result['sequence_id']}")
    print(f"\nFolder: {result['folder_url']}")
    print(f"Email Sequences: {result['doc_url']}")

    print("\nSEQUENCE STRUCTURE:")
    print("  3 sequences (A/B/C):")
    print("    - Sequence A: Email 1, Email 2, Email 3 (Problem-focused)")
    print("    - Sequence B: Email 1, Email 2, Email 3 (Context-add)")
    print("    - Sequence C: Email 1, Email 2, Email 3 (Curiosity-driven)")
    print("  Total: 9 email templates")

    print("\nKEY FEATURES:")
    print("  ✓ Niche hardcoded in email copy (not a placeholder)")
    print("  ✓ {first_name} and {company_name} remain as placeholders")
    print("  ✓ Generic templates ready for any lead list")
    print("  ✓ Ready for spintax addition with /add-spintax")

    print("\nNEXT STEPS:")
    print("  1. Review email sequences in Google Doc")
    print("  2. Edit sequences if needed")
    print("  3. Run /add-spintax <doc_id> when ready")
    print("  4. Use with any lead list in SmartLead upload")

    print("=" * 60)


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python generate_cold_email_copy.py <niche_name>")
        print("  python generate_cold_email_copy.py <excel_file>")
        print("\nExamples:")
        print("  python generate_cold_email_copy.py Marketing")
        print("  python generate_cold_email_copy.py SaaS")
        print("  python generate_cold_email_copy.py leads_sendable.xlsx")
        sys.exit(1)

    input_arg = sys.argv[1]

    # Check if input is an Excel file or a niche name
    input_path = Path(input_arg)

    if input_path.exists() and input_path.suffix.lower() in ['.xlsx', '.xls']:
        # Excel file mode (original workflow)
        result = process_excel_file(input_arg)
        print_summary(result)
    else:
        # Niche name mode (new workflow)
        result = generate_sequences_for_niche(input_arg)
        print_niche_summary(result)


if __name__ == '__main__':
    main()
