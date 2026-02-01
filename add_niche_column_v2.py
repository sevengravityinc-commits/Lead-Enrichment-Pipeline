"""
Add Niche Column V2 - Fast Batch AI Categorization
Categorizes companies into user-specified niches using intelligent batch AI.
Uses all available context columns: Company Name (required), Title, Keywords, Industry, Sub Industry (optional).
"""

import os
import sys
import json
import time
import csv
import argparse
import requests
from pathlib import Path
from typing import List, Dict, Optional
from collections import Counter, defaultdict
from dotenv import load_dotenv

load_dotenv()

try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Configuration
DEFAULT_BATCH_SIZE = 50
DEFAULT_MODEL = "openai/gpt-4o-mini"
DEFAULT_COLUMN_NAME = "Verified_Niche"
API_DELAY = 0.5  # Seconds between API calls
MAX_RETRIES = 3  # Number of retry attempts for invalid responses

OPENROUTER_API_KEY = os.getenv('OPENROUTER_API_KEY')


def create_classification_prompt(categories: List[str]) -> str:
    """Generate classification prompt with user-specified categories."""
    categories_str = ", ".join(categories)

    return f"""# ROLE
You are a B2B market research analyst specializing in lead segmentation for targeted outreach campaigns.

# TASK CONTEXT
You're helping a sales team segment 10,000+ leads into campaign-specific lists. Each category will receive customized messaging, so accuracy is critical. Misclassification = wasted outreach + damaged reputation.

# YOUR GOAL
Classify each company into exactly ONE category: {categories_str}

# CLASSIFICATION CRITERIA

**What makes a good classification:**
- Based on PRIMARY revenue source (not side projects)
- Uses ACTUAL business model (not aspirations in job titles)
- Considers company size/stage (startup SaaS ≠ enterprise SaaS)
- Prioritizes concrete data (keywords, industry) over vague signals (job title)

**Quality standards:**
- 95%+ accuracy expected
- When uncertain, choose the MOST SPECIFIC category
- Example: "Marketing Agency" > "Agency" > "Services"

# DECISION FRAMEWORK

**Priority order for classification:**
1. Keywords (if provided) - direct business indicators
2. Industry + Sub Industry - broad category context
3. Company Name - often contains business model hints
4. Job Title - weakest signal, use only as tiebreaker

**Example logic:**
- Keywords: "SaaS, B2B, Sales Software" → SaaS (high confidence)
- Industry: "Technology", Job Title: "Marketing Director" → Could be SaaS or Agency
  → Check company name for "Software", "Tech", "Platform" → SaaS
  → If name unclear → Default to Technology

# EDGE CASES

- **Holding companies** → Classify by primary operating subsidiary
- **Conglomerates** → Use most recent/dominant business line
- **Pivoting startups** → Use current focus (not historical)
- **Insufficient data** → Mark as "Insufficient Data" (never guess)

# OUTPUT FORMAT
Return ONLY a JSON array of categories in the same order as input. No explanations, no markdown.

Example input: ["Acme Corp | CEO | SaaS, B2B | Technology", "Blue Ocean Agency | Marketing Director | ..."]
Example output: ["{categories[0]}", "{categories[1] if len(categories) > 1 else categories[0]}"]

# VALIDATION
Every category MUST be from this exact list: {categories_str}
Invalid categories will be rejected and retried (up to 3 attempts)."""


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns (case-insensitive, partial matching)."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx
    return None


def batch_categorize(companies: List[Dict], categories: List[str], model: str) -> List[str]:
    """
    Send a batch of companies to OpenRouter and get categories back.
    Includes retry logic to ensure accurate categorization.

    Args:
        companies: List of dicts with 'name' and optional context keys
        categories: List of valid category names
        model: AI model to use

    Returns:
        List of categories matching input length
    """
    # Format companies for prompt - include all available context
    company_lines = []
    for c in companies:
        parts = [f"Company: {c.get('name', 'Unknown')}"]

        if c.get('title'):
            parts.append(f"Title: {c['title']}")
        if c.get('keywords'):
            parts.append(f"Keywords: {c['keywords']}")
        if c.get('industry'):
            parts.append(f"Industry: {c['industry']}")
        if c.get('sub_industry'):
            parts.append(f"Sub Industry: {c['sub_industry']}")

        company_lines.append(" | ".join(parts))

    companies_json = json.dumps(company_lines)
    prompt = create_classification_prompt(categories)

    # Retry loop (up to MAX_RETRIES attempts)
    for attempt in range(MAX_RETRIES):
        try:
            response = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": model,
                    "max_tokens": len(companies) * 20,
                    "messages": [
                        {"role": "system", "content": prompt},
                        {"role": "user", "content": f"Classify these companies:\n{companies_json}"}
                    ]
                },
                timeout=60
            )

            response.raise_for_status()
            data = response.json()
            response_text = data['choices'][0]['message']['content'].strip()

            # Parse JSON response (handle markdown code blocks)
            if response_text.startswith("```"):
                response_text = response_text.split("```")[1]
                if response_text.startswith("json"):
                    response_text = response_text[4:]
                response_text = response_text.strip()

            categories_result = json.loads(response_text)

            # Validate all categories
            valid_categories_set = set(categories)
            validated = []
            all_valid = True

            for cat in categories_result:
                if cat in valid_categories_set:
                    validated.append(cat)
                else:
                    all_valid = False
                    break

            # Check if we got the right number of results
            if all_valid and len(validated) == len(companies):
                return validated
            else:
                # Invalid response - retry
                if attempt < MAX_RETRIES - 1:
                    print(f"  Warning: Invalid categories in response, retrying (attempt {attempt + 2}/{MAX_RETRIES})...")
                    time.sleep(1)  # Brief delay before retry
                    continue
                else:
                    # All retries failed
                    print(f"  Error: Failed to get valid categories after {MAX_RETRIES} attempts")
                    return ["Categorization Failed"] * len(companies)

        except json.JSONDecodeError as e:
            if attempt < MAX_RETRIES - 1:
                print(f"  Warning: JSON parse error, retrying (attempt {attempt + 2}/{MAX_RETRIES})...")
                time.sleep(1)
                continue
            else:
                print(f"  Error: JSON parse failed after {MAX_RETRIES} attempts: {e}")
                return ["Categorization Failed"] * len(companies)

        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                print(f"  Warning: API error, retrying (attempt {attempt + 2}/{MAX_RETRIES})...")
                time.sleep(1)
                continue
            else:
                print(f"  Error: API failed after {MAX_RETRIES} attempts: {e}")
                return ["Categorization Failed"] * len(companies)

    # Should never reach here, but safety fallback
    return ["Categorization Failed"] * len(companies)


def create_invalid_data_file(file_path: Path, ws, headers: List[str], niche_col_idx: int, valid_categories: List[str]) -> tuple:
    """
    Create separate Excel file with all invalid/failed rows.

    Returns:
        (invalid_file_path, invalid_count)
    """
    # Create new workbook for invalid data
    invalid_wb = Workbook()
    invalid_ws = invalid_wb.active

    # Copy headers
    for col_idx, header in enumerate(headers, 1):
        invalid_ws.cell(row=1, column=col_idx, value=header)

    # Define invalid categories
    invalid_categories = {
        "Categorization Failed",
        "Insufficient Data",
        "Research Required",
        None,
        ""
    }

    # Find all invalid rows
    invalid_row_count = 0
    for row_idx in range(2, ws.max_row + 1):
        niche = ws.cell(row=row_idx, column=niche_col_idx + 1).value

        # Check if invalid
        is_invalid = (
            niche in invalid_categories or
            (niche and niche.strip() == "") or
            (niche and niche not in valid_categories)
        )

        if is_invalid:
            invalid_row_count += 1
            # Copy entire row
            for col_idx in range(1, len(headers) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                invalid_ws.cell(row=invalid_row_count + 1, column=col_idx, value=value)

    # Save invalid data file (only if there are invalid rows)
    if invalid_row_count > 0:
        base_name = file_path.stem
        invalid_filename = f"{base_name}_INVALID.xlsx"
        invalid_path = file_path.parent / invalid_filename

        invalid_wb.save(invalid_path)
        print(f"\nInvalid data file created: {invalid_filename} ({invalid_row_count} rows)")
        print(f"Please review these rows manually and re-run the script to process them.")
        return invalid_path, invalid_row_count
    else:
        print("\nNo invalid rows found - all data categorized successfully!")
        return None, 0


def split_to_csv_by_niche(file_path: Path, ws, headers: List[str], niche_col_idx: int, valid_categories: List[str]) -> Path:
    """
    Split Excel file into CSV files by Verified_Niche value.
    Only includes valid categories.

    Returns:
        Path to output folder
    """
    # Create output folder
    base_name = file_path.stem
    output_folder = file_path.parent / f"{base_name}_by_niche"
    output_folder.mkdir(exist_ok=True)

    # Group rows by niche
    niche_groups = defaultdict(list)

    for row_idx in range(2, ws.max_row + 1):
        niche = ws.cell(row=row_idx, column=niche_col_idx + 1).value

        # Only include valid categories
        if niche and niche in valid_categories:
            row_data = [ws.cell(row=row_idx, column=col).value
                       for col in range(1, len(headers) + 1)]
            niche_groups[niche].append(row_data)

    # Create CSV for each niche
    print(f"\nSplitting into CSV files by niche...")
    for niche in valid_categories:  # Maintain order from user input
        if niche not in niche_groups:
            continue

        rows = niche_groups[niche]

        # Clean niche name for filename
        safe_niche = niche.replace('/', '-').replace('\\', '-').replace('|', '-')
        csv_filename = f"{base_name}_{safe_niche}.csv"
        csv_path = output_folder / csv_filename

        # Write CSV
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

        print(f"  Created: {csv_filename} ({len(rows)} rows)")

    return output_folder


def process_excel_file(
    file_path: str,
    categories: List[str],
    batch_size: int = DEFAULT_BATCH_SIZE,
    model: str = DEFAULT_MODEL,
    column_name: str = DEFAULT_COLUMN_NAME,
    verbose: bool = True
) -> Dict:
    """
    Process Excel file and add niche column using batch AI categorization.

    Args:
        file_path: Path to Excel file
        categories: List of valid category names
        batch_size: Companies per API call
        model: AI model to use
        column_name: Name of output column
        verbose: Whether to print progress

    Returns:
        Dict with processing statistics
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    if not OPENROUTER_API_KEY:
        return {"error": "OPENROUTER_API_KEY not set in environment"}

    if not categories or len(categories) < 2:
        return {"error": "Must provide at least 2 categories"}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    # Create backup
    backup_path = file_path.parent / f"{file_path.stem}_backup{file_path.suffix}"
    if not backup_path.exists():
        import shutil
        shutil.copy2(file_path, backup_path)
        if verbose:
            print(f"Backup created: {backup_path.name}")

    # Load workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {"error": f"Failed to load Excel file: {e}"}

    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Intelligent column detection
    columns = {
        'company': find_column_index(headers, ['company', 'company_name', 'company name', 'name']),
        'title': find_column_index(headers, ['title', 'job_title', 'job title', 'position']),
        'keywords': find_column_index(headers, ['keywords', 'keyword', 'tags']),
        'industry': find_column_index(headers, ['industry', 'sector', 'vertical']),
        'sub_industry': find_column_index(headers, ['sub_industry', 'sub industry', 'subcategory', 'sub_category'])
    }

    if columns['company'] is None:
        return {"error": "Could not find Company Name column"}

    # Report detected columns
    if verbose:
        print(f"\nDetected columns:")
        print(f"  Company Name: column {columns['company'] + 1} (required)")
        for key in ['title', 'keywords', 'industry', 'sub_industry']:
            if columns[key] is not None:
                print(f"  {key.replace('_', ' ').title()}: column {columns[key] + 1} (optional context)")

    # Find or create niche column
    niche_col = find_column_index(headers, [column_name.lower()])
    if niche_col is None:
        niche_col = len(headers)
        ws.cell(row=1, column=niche_col + 1, value=column_name)
        headers.append(column_name)

    # Get total rows
    total_rows = ws.max_row - 1

    if verbose:
        print(f"\nProcessing {total_rows} rows in batches of {batch_size}...")
        print(f"Estimated API calls: {(total_rows + batch_size - 1) // batch_size}")
        print(f"Model: {model}")
        print(f"Categories: {', '.join(categories)}")
        print()

    # Statistics
    valid_categories_set = set(categories)
    stats = {
        "total_rows": total_rows,
        "processed": 0,
        "skipped": 0,
        "batches": 0,
        "distribution": Counter()
    }

    # Process in batches
    batch = []
    batch_rows = []

    for row_idx in range(2, ws.max_row + 1):
        # Check if already categorized with valid category
        existing = ws.cell(row=row_idx, column=niche_col + 1).value
        if existing and existing.strip() in valid_categories_set:
            stats["skipped"] += 1
            stats["distribution"][existing.strip()] += 1
            continue

        # Get company data
        company_name = ws.cell(row=row_idx, column=columns['company'] + 1).value

        if not company_name:
            ws.cell(row=row_idx, column=niche_col + 1, value='Insufficient Data')
            stats["processed"] += 1
            stats["distribution"]["Insufficient Data"] += 1
            continue

        # Build company dict with all available context
        company_data = {'name': str(company_name).strip()}

        for key in ['title', 'keywords', 'industry', 'sub_industry']:
            if columns[key] is not None:
                value = ws.cell(row=row_idx, column=columns[key] + 1).value
                if value:
                    company_data[key] = str(value).strip()

        batch.append(company_data)
        batch_rows.append(row_idx)

        # Process batch when full
        if len(batch) >= batch_size:
            stats["batches"] += 1

            if verbose:
                progress = (stats["processed"] + stats["skipped"]) / total_rows * 100
                print(f"  Batch {stats['batches']}: rows {batch_rows[0]}-{batch_rows[-1]} ({progress:.1f}% complete)")

            # Get categories from AI
            results = batch_categorize(batch, categories, model)

            # Write results
            for row, category in zip(batch_rows, results):
                ws.cell(row=row, column=niche_col + 1, value=category)
                stats["distribution"][category] += 1
                stats["processed"] += 1

            # Save progress
            try:
                wb.save(file_path)
            except Exception as e:
                print(f"  Warning: Could not save progress: {e}")

            # Clear batch
            batch = []
            batch_rows = []

            # Rate limiting
            time.sleep(API_DELAY)

    # Process remaining batch
    if batch:
        stats["batches"] += 1

        if verbose:
            print(f"  Batch {stats['batches']} (final): rows {batch_rows[0]}-{batch_rows[-1]}")

        results = batch_categorize(batch, categories, model)

        for row, category in zip(batch_rows, results):
            ws.cell(row=row, column=niche_col + 1, value=category)
            stats["distribution"][category] += 1
            stats["processed"] += 1

    # Final save
    try:
        wb.save(file_path)
        if verbose:
            print(f"\nSaved: {file_path}")
    except Exception as e:
        wb.close()
        return {"error": f"Failed to save: {e}", **stats}

    # Create invalid data file
    invalid_path, invalid_count = create_invalid_data_file(
        file_path, ws, headers, niche_col, valid_categories_set
    )

    # Split to CSV files
    output_folder = split_to_csv_by_niche(
        file_path, ws, headers, niche_col, valid_categories_set
    )

    wb.close()

    stats['invalid_count'] = invalid_count
    stats['invalid_file'] = str(invalid_path) if invalid_path else None
    stats['csv_folder'] = str(output_folder)

    return stats


def main():
    parser = argparse.ArgumentParser(
        description='Fast batch AI categorization of companies into user-specified niches'
    )
    parser.add_argument('file', help='Path to Excel file')
    parser.add_argument(
        '--categories',
        required=True,
        help='Pipe-separated list of categories (e.g., "Marketing|Advertising|PR")'
    )
    parser.add_argument(
        '--batch-size',
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help=f'Companies per API call (default: {DEFAULT_BATCH_SIZE})'
    )
    parser.add_argument(
        '--model',
        default=DEFAULT_MODEL,
        help=f'AI model to use (default: {DEFAULT_MODEL})'
    )
    parser.add_argument(
        '--column-name',
        default=DEFAULT_COLUMN_NAME,
        help=f'Output column name (default: {DEFAULT_COLUMN_NAME})'
    )

    args = parser.parse_args()

    # Parse categories
    categories = [c.strip() for c in args.categories.split('|') if c.strip()]

    if len(categories) < 2:
        print("Error: Must provide at least 2 categories")
        sys.exit(1)

    print(f"\n{'='*60}")
    print("FAST BATCH NICHE CATEGORIZATION")
    print(f"{'='*60}")
    print(f"File: {args.file}")
    print(f"Categories: {', '.join(categories)}")
    print(f"Batch size: {args.batch_size}")
    print(f"Model: {args.model}")

    start_time = time.time()

    result = process_excel_file(
        args.file,
        categories,
        batch_size=args.batch_size,
        model=args.model,
        column_name=args.column_name,
        verbose=True
    )

    elapsed = time.time() - start_time

    print(f"\n{'='*60}")
    print("RESULTS")
    print(f"{'='*60}")

    if "error" in result:
        print(f"Error: {result['error']}")
        sys.exit(1)

    print(f"Total rows: {result['total_rows']}")
    print(f"Processed: {result['processed']}")
    print(f"Skipped (already done): {result['skipped']}")
    print(f"API batches: {result['batches']}")
    print(f"Time elapsed: {elapsed:.1f} seconds")

    # Separate valid and invalid categories
    valid_dist = {k: v for k, v in result['distribution'].items() if k in categories}
    invalid_dist = {k: v for k, v in result['distribution'].items() if k not in categories}

    if valid_dist:
        print(f"\nValid Category Distribution:")
        for category in categories:  # Maintain user's order
            if category in valid_dist:
                count = valid_dist[category]
                pct = count / result['total_rows'] * 100 if result['total_rows'] > 0 else 0
                print(f"  {category}: {count} ({pct:.1f}%)")

    if invalid_dist:
        print(f"\nInvalid/Failed (in {Path(args.file).stem}_INVALID.xlsx):")
        for category, count in sorted(invalid_dist.items()):
            pct = count / result['total_rows'] * 100 if result['total_rows'] > 0 else 0
            print(f"  {category}: {count} ({pct:.1f}%)")

    print(f"\nCSV files saved to: {result['csv_folder']}")


if __name__ == '__main__':
    main()
