"""
Add Clean Company Column Script
Main orchestrator that processes Excel files and adds a Clean_Company_Name column
using AI-powered normalization.
"""

import os
import sys
import time
import shutil
from pathlib import Path
from typing import Optional, List, Dict
from collections import Counter

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from normalize_company_name import normalize_company_name, normalize_batch, BATCH_SIZE


# Column name patterns for finding company name
COMPANY_COLUMN_PATTERNS = [
    'company', 'company_name', 'companyname', 'name', 'organization',
    'org', 'business', 'business_name', 'firm', 'company name',
    'organization_name', 'org_name', 'account', 'account_name'
]

# Rate limiting between batch API calls
BATCH_DELAY = 1.0  # seconds between batch requests


def find_column_index(headers: List[str], patterns: List[str]) -> Optional[int]:
    """Find column index matching any of the patterns."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]

    for pattern in patterns:
        pattern_lower = pattern.lower()
        for idx, header in enumerate(headers_lower):
            if pattern_lower == header or pattern_lower in header:
                return idx

    return None


def get_headers(ws) -> List[str]:
    """Get header row from worksheet."""
    return [cell.value for cell in ws[1]]


def process_excel_file(file_path: str, verbose: bool = True) -> Dict:
    """
    Process a single Excel file and add Clean_Company_Name column.
    Uses batch processing for efficiency (50 names per API call).

    For 20,000 rows: ~400 API calls instead of 20,000.

    Args:
        file_path: Path to the Excel file
        verbose: Whether to print progress

    Returns:
        Dict with processing statistics
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": f"File not found: {file_path}"}

    if not file_path.suffix.lower() in ['.xlsx', '.xls']:
        return {"error": f"Not an Excel file: {file_path}"}

    # Create backup
    backup_path = file_path.parent / f"{file_path.stem}_backup{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    if verbose:
        print(f"Backup created: {backup_path}")

    # Load workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {"error": f"Failed to load Excel file: {str(e)}"}

    stats = {
        "file": str(file_path),
        "sheets_processed": 0,
        "rows_processed": 0,
        "successful_normalizations": 0,
        "failed_normalizations": 0,
        "unchanged": 0,
        "samples": []
    }

    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.max_row < 2:  # Empty or header only
            continue

        headers = get_headers(ws)
        if not headers or all(h is None for h in headers):
            continue

        # Find company name column
        company_col_idx = find_column_index(headers, COMPANY_COLUMN_PATTERNS)

        if company_col_idx is None:
            if verbose:
                print(f"  Sheet '{sheet_name}': No company column found, skipping")
            continue

        # Find or create Clean_Company_Name column
        clean_col_idx = find_column_index(headers, ['clean_company_name', 'clean company name', 'normalized_company'])

        if clean_col_idx is None:
            # Add new column
            clean_col_idx = len(headers)
            ws.cell(row=1, column=clean_col_idx + 1, value='Clean_Company_Name')

        if verbose:
            print(f"  Processing sheet '{sheet_name}' ({ws.max_row - 1} rows)...")

        stats["sheets_processed"] += 1

        # PHASE 1: Collect all company names and their row positions
        company_data = []  # List of (row_idx, company_name)
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=company_col_idx + 1).value
            if cell_value:
                company_name = str(cell_value).strip()
                if company_name:
                    company_data.append((row_idx, company_name))
                else:
                    ws.cell(row=row_idx, column=clean_col_idx + 1, value='')
            else:
                ws.cell(row=row_idx, column=clean_col_idx + 1, value='')

        if not company_data:
            continue

        if verbose:
            print(f"  Collected {len(company_data)} company names for batch processing...")

        # PHASE 2: Batch normalize all names
        company_names = [name for _, name in company_data]
        results = normalize_batch(company_names, delay=BATCH_DELAY)

        # PHASE 3: Write results back to spreadsheet
        for i, (row_idx, original_name) in enumerate(company_data):
            result = results[i]
            stats["rows_processed"] += 1

            if result.success:
                ws.cell(row=row_idx, column=clean_col_idx + 1, value=result.normalized)
                stats["successful_normalizations"] += 1

                # Track if actually changed
                if result.normalized != result.original:
                    if len(stats["samples"]) < 10:
                        stats["samples"].append({
                            "original": result.original,
                            "normalized": result.normalized
                        })
                else:
                    stats["unchanged"] += 1
            else:
                # On failure, keep original
                ws.cell(row=row_idx, column=clean_col_idx + 1, value=original_name)
                stats["failed_normalizations"] += 1

    # Save workbook
    try:
        wb.save(file_path)
        if verbose:
            print(f"Saved: {file_path}")
    except Exception as e:
        return {"error": f"Failed to save: {str(e)}", **stats}

    return stats


def process_folder(folder_path: str, verbose: bool = True) -> List[Dict]:
    """
    Process all Excel files in a folder recursively.

    Args:
        folder_path: Path to the folder
        verbose: Whether to print progress

    Returns:
        List of stats for each file processed
    """
    folder_path = Path(folder_path)
    if not folder_path.exists():
        return [{"error": f"Folder not found: {folder_path}"}]

    if not folder_path.is_dir():
        return [{"error": f"Not a folder: {folder_path}"}]

    # Find all Excel files
    excel_files = list(folder_path.glob('**/*.xlsx')) + list(folder_path.glob('**/*.xls'))

    # Filter out backup files
    excel_files = [f for f in excel_files if '_backup' not in f.stem]

    if not excel_files:
        return [{"error": f"No Excel files found in: {folder_path}"}]

    if verbose:
        print(f"Found {len(excel_files)} Excel files to process")

    results = []
    for file_path in excel_files:
        if verbose:
            print(f"\nProcessing: {file_path.name}")
        result = process_excel_file(str(file_path), verbose=verbose)
        results.append(result)

    return results


def print_summary(results: List[Dict]):
    """Print summary of processing results."""
    print("\n" + "=" * 60)
    print("PROCESSING SUMMARY")
    print("=" * 60)
    print(f"  Model: {os.getenv('NORMALIZER_MODEL', 'deepseek/deepseek-chat')}")
    print(f"  Batch size: {BATCH_SIZE} names per API call")

    total_rows = 0
    total_success = 0
    total_failed = 0
    total_unchanged = 0
    all_samples = []

    for result in results:
        if "error" in result and result.get("rows_processed", 0) == 0:
            print(f"ERROR: {result['error']}")
        else:
            if "file" in result:
                print(f"\nFile: {Path(result['file']).name}")
            if "sheets_processed" in result:
                print(f"  Sheets processed: {result['sheets_processed']}")
            if "rows_processed" in result:
                print(f"  Rows processed: {result['rows_processed']}")
                total_rows += result['rows_processed']
            if "successful_normalizations" in result:
                print(f"  Successful: {result['successful_normalizations']}")
                total_success += result['successful_normalizations']
            if "failed_normalizations" in result:
                print(f"  Failed: {result['failed_normalizations']}")
                total_failed += result['failed_normalizations']
            if "unchanged" in result:
                print(f"  Unchanged: {result['unchanged']}")
                total_unchanged += result['unchanged']
            if "samples" in result:
                all_samples.extend(result['samples'])

    print("\n" + "-" * 60)
    print("TOTALS")
    print(f"  Total rows processed: {total_rows}")
    print(f"  Total successful: {total_success}")
    print(f"  Total failed: {total_failed}")
    print(f"  Total unchanged: {total_unchanged}")
    print(f"  Total changed: {total_success - total_unchanged}")

    if all_samples:
        print("\nSAMPLE NORMALIZATIONS:")
        for sample in all_samples[:10]:
            print(f"  \"{sample['original']}\" -> \"{sample['normalized']}\"")

    print("=" * 60)


def main():
    if len(sys.argv) < 2:
        print("Usage: python add_clean_company_column.py <file_or_folder_path>")
        print("\nExamples:")
        print("  python add_clean_company_column.py companies.xlsx")
        print("  python add_clean_company_column.py ./leads_folder/")
        sys.exit(1)

    path = sys.argv[1]

    if Path(path).is_file():
        result = process_excel_file(path)
        print_summary([result])
    elif Path(path).is_dir():
        results = process_folder(path)
        print_summary(results)
    else:
        print(f"Error: Path not found: {path}")
        sys.exit(1)


if __name__ == '__main__':
    main()
